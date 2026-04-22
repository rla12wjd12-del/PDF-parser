#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF 업로드(선택) → 순수 추출 표(extract_tables_merged) → Excel 다운로드(정제 없음)

실행:
  streamlit run scripts/export_raw_tables_uploader_app.py
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from parsers.table_settings import extract_tables_merged  # noqa: E402


def _cell(v: object) -> str:
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").replace("\r", "\n")


def _autofit(ws, max_w: float = 55.0) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        m = 8
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                m = min(max(m, len(str(v)) + 1), int(max_w))
        ws.column_dimensions[letter].width = float(m)


def build_raw_tables_xlsx(pdf_bytes: bytes, *, pages: list[int] | None = None) -> bytes:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "INDEX"
    ws0.append(["page_idx", "page_no", "table_count"])

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        total = len(pdf.pages)
        if pages is None:
            pages = list(range(total))
        pages = [p for p in pages if 0 <= p < total]

        for pi in pages:
            page = pdf.pages[pi]
            tables = extract_tables_merged(page)
            ws0.append([pi, pi + 1, len(tables)])

            ws = wb.create_sheet(f"p{pi + 1:03d}")
            ws.cell(row=1, column=1, value=f"page_idx={pi} (page_no={pi + 1}) — RAW extract_tables_merged")
            row = 3
            for ti, t in enumerate(tables):
                if not t:
                    continue
                max_cols = max((len(r) for r in t if r), default=0)
                ws.cell(row=row, column=1, value=f"[TABLE] RAW {ti} ({len(t)} x {max_cols})")
                row += 1
                for tr in t:
                    for cj, c in enumerate(tr or []):
                        ws.cell(row=row, column=cj + 1, value=_cell(c))
                    row += 1
                row += 2
            _autofit(ws)

    _autofit(ws0)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_pages_spec(spec: str) -> list[int] | None:
    s = (spec or "").strip()
    if not s:
        return None
    # "0-5" or "0 1 2"
    if "-" in s and all(x.strip().isdigit() for x in s.split("-", 1)):
        a, b = [int(x.strip()) for x in s.split("-", 1)]
        if a > b:
            a, b = b, a
        return list(range(a, b + 1))
    out: list[int] = []
    for tok in s.replace(",", " ").split():
        if tok.isdigit():
            out.append(int(tok))
    return sorted(set(out)) if out else None


def main() -> None:
    st.set_page_config(page_title="RAW PDF 표 추출 → Excel", layout="wide")
    st.title("RAW PDF 표 추출 → Excel (정제 없음)")
    st.caption("pdfplumber `extract_tables_merged` 결과를 그대로 Excel로 내려받습니다.")

    up = st.file_uploader("PDF 파일 업로드", type=["pdf"])
    pages_spec = st.text_input("페이지(0부터) 선택 (예: 0-2 또는 0 1 2, 비우면 전체)", value="")

    if up is None:
        st.info("PDF를 업로드하면 RAW 표 추출 결과를 Excel로 다운로드할 수 있습니다.")
        return

    data = up.getvalue()
    if not data:
        st.error("업로드된 파일이 비어있습니다.")
        return

    pages = parse_pages_spec(pages_spec)
    try:
        xlsx_bytes = build_raw_tables_xlsx(data, pages=pages)
    except Exception as e:
        st.error("표 추출/엑셀 생성 실패")
        st.exception(e)
        return

    stem = Path(up.name).stem
    out_name = f"{stem}_raw_tables.xlsx"
    st.download_button(
        label="Excel 다운로드",
        data=xlsx_bytes,
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.success("생성 완료. 위 버튼으로 다운로드하세요.")


if __name__ == "__main__":
    main()

