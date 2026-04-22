#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
JSON/필드 매핑 이전 단계에서 pdfplumber가 인식한 표(셀 그리드)를 Excel로 덤프한다.

운영 파서와 동일한 설정을 쓰려면 `extract_tables_merged`(가상선+lines → text → 기본값)를 사용한다.
전략별로만 보고 싶으면 --by-strategy 를 지정한다.

사용 예:
  python scripts/export_recognized_tables_to_excel.py "testpdf/sample.pdf"
  python scripts/export_recognized_tables_to_excel.py "testpdf/sample.pdf" -o _out/tables_preview.xlsx --pages 0-2
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

# 프로젝트 루트를 path에 추가
_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import pdfplumber
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from parsers.table_settings import (
    LINE_TABLE_SETTINGS,
    TEXT_TABLE_SETTINGS,
    extract_tables_merged,
    safe_extract_tables,
)

_HEADER_KEYWORDS = [
    "사업명",
    "발주자",
    "공사(용역)개요",
    "직무분야",
    "전문분야",
    "직위",
    "책임정도",
    "공사(용역)금액",
    "적용 공법",
    "적용 신기술",
    "시설물 종류",
    "비고",
]

# 페이지 상/하단 기본 문구/헤더/푸터 제거용(표 내부에 섞여 들어오는 케이스)
_BOILERPLATE_RE = re.compile(
    r"(?i)\bPage\s*:\s*\d+\s*/\s*\d+\b|"
    r"\bPage\s*\d+\s*/\s*\d+\b|"
    r"문서\s*확인\s*번호|발급\s*번호|관리\s*번호|"
    r"본\s*증명서는|진위\s*확인|홈페이지|"
    r"건설기술\s*진흥법|시행령|제\s*45조|"
    r"^\s*\(\s*\d+쪽\s*중\s*\d+쪽\s*\)\s*$"
)


def _row_text(row: list[object]) -> str:
    s = " ".join(_cell_str(c) for c in (row or []) if _cell_str(c).strip())
    return re.sub(r"\s+", " ", s).strip()


def _is_header_like_row(row: list[object]) -> bool:
    t = _row_text(row)
    if not t:
        return False
    hit = sum(1 for k in _HEADER_KEYWORDS if k in t)
    # 헤더는 키워드가 여러 개 묶여 등장
    if hit >= 2:
        return True
    # 좌측 제목 라벨
    if re.search(r"\b[12]\.\s*(기술경력|건설사업관리|감리경력)\b", t):
        return True
    return False


def _is_boilerplate_row(row: list[object]) -> bool:
    t = _row_text(row)
    if not t:
        return False
    if _BOILERPLATE_RE.search(t):
        return True
    # 'Pag e : 9 / 31' 처럼 깨진 케이스 보강
    t2 = t.replace(" ", "")
    if "Page:" in t2 and "/" in t2:
        return True
    return False


def clean_table(table: list[list[object]]) -> list[list[object]]:
    """
    원시 표에서 헤더/푸터/기본 문구를 제거한 '정제 표'를 만든다.
    - 하드코딩된 행 번호가 아니라, 행 텍스트 신호로 판단
    """
    if not table:
        return []
    out: list[list[object]] = []
    header_seen = 0
    for r in table:
        if not r or not _row_text(r):
            # 연속 공백 라인은 1개만 유지
            if out and _row_text(out[-1]):
                out.append([""])
            continue
        if _is_boilerplate_row(r):
            continue
        if _is_header_like_row(r):
            # 헤더는 최대 2~3줄까지 허용(그 이후 반복 헤더는 제거)
            header_seen += 1
            if header_seen <= 3:
                out.append(r)
            continue
        out.append(r)
    # trailing blank 제거
    while out and not _row_text(out[-1]):
        out.pop()
    return out

def _table_join_text(table: list[list[object]], *, max_rows: int = 20) -> str:
    parts: list[str] = []
    for row in (table or [])[:max_rows]:
        if not row:
            continue
        for c in row:
            s = _cell_str(c).strip()
            if s:
                parts.append(s)
    return " ".join(parts)


def _is_section_title_or_note_table(table: list[list[object]]) -> bool:
    """
    '1. 기술경력', '2. 건설사업관리 및 감리경력', '※ ...' 같은 제목/주석 중심 블록을 배제한다.
    (하드코딩된 위치가 아니라, 텍스트 신호로 판정)
    """
    t = re.sub(r"\s+", " ", _table_join_text(table, max_rows=12)).strip()
    if not t:
        return True
    if "※" in t:
        return True
    if re.search(r"\b[12]\.\s*(기술경력|건설사업관리|감리경력)\b", t):
        return True
    if "책임정도의" in t and "보정계수" in t:
        return True
    # 제목/설명만 있고 헤더 키워드가 거의 없는 경우
    hit = sum(1 for k in _HEADER_KEYWORDS if k in t)
    if hit <= 1 and len(t) >= 80:
        return True
    return False


def _table_score_best_only(table: list[list[object]]) -> tuple[int, int, int, int]:
    """
    엑셀 덤프에서 페이지당 '최적 표 1개'를 고르기 위한 스코어.
    - 헤더 키워드가 많을수록 가산점
    - 제목/주석 블록이면 강한 감점
    - 열 수/의미 있는 행 수로 형태 점수 보강
    """
    if not table:
        return (-10_000, 0, 0, 0)
    max_cols = max((len(r) for r in table if r), default=0)
    if max_cols < 3:
        return (-10_000, max_cols, 0, len(table))
    joined = _table_join_text(table, max_rows=18)
    hit = sum(1 for k in _HEADER_KEYWORDS if k in joined)
    has_left = "참여기간" in joined
    has_right = "비고" in joined
    # 데이터 행 수(비어있지 않은 행)
    non_empty_rows = 0
    for r in table:
        if any(_cell_str(c).strip() for c in (r or [])):
            non_empty_rows += 1
    base = hit * 220 + max_cols * 25 + non_empty_rows
    # 핵심 컬럼(좌: 참여기간, 우: 비고)이 함께 있으면 강한 가산점
    if has_left and has_right:
        base += 800
    # 헤더는 풍부한데 열 수가 너무 적으면(경계선 누락) 감점
    if hit >= 3 and max_cols <= 4:
        base -= 900
    if _is_section_title_or_note_table(table):
        base -= 2000
    return (base, hit, max_cols, non_empty_rows)


def _pick_best_table_only(tables: list[list[list[object]]]) -> list[list[list[object]]]:
    best = None
    best_sc = None
    for t in tables or []:
        sc = _table_score_best_only(t)
        if best_sc is None or sc > best_sc:
            best_sc = sc
            best = t
    return [best] if best else []


def _parse_page_range(spec: str | None, total: int) -> list[int]:
    if not spec or not spec.strip():
        return list(range(total))
    spec = spec.strip()
    if re.fullmatch(r"\d+-\d+", spec):
        a, b = map(int, spec.split("-", 1))
        return [i for i in range(total) if a <= i <= b]
    parts: list[int] = []
    for tok in spec.replace(",", " ").split():
        if tok.isdigit():
            parts.append(int(tok))
    return sorted({p for p in parts if 0 <= p < total})


def _cell_str(val: object) -> str:
    if val is None:
        return ""
    s = str(val).replace("\r\n", "\n").replace("\r", "\n")
    return s


def _safe_sheet_name(used: set[str], name: str) -> str:
    s = re.sub(r'[\[\]:*?/\\]', "_", name)[:31] or "Sheet"
    base = s
    n = 1
    while s in used:
        suf = f"_{n}"
        s = (base[: max(0, 31 - len(suf))] + suf).strip() or f"sh_{n}"
        n += 1
    used.add(s)
    return s


def _infer_header_start_row(table: list[list[object]]) -> int | None:
    if not table:
        return None
    date_like = re.compile(r"^\s*\d{4}\.\d{2}\.\d{2}\b")
    for i, r in enumerate(table[:12]):
        if i == 0:
            continue
        if not r:
            continue
        non_empty = [str(c).strip() for c in r if str(c or "").strip()]
        if len(non_empty) < 3:
            continue
        c0 = str(r[0] or "").strip() if len(r) >= 1 else ""
        if c0 and date_like.match(c0):
            continue
        return i
    return None


def _header_override_rows(ncols: int) -> dict[int, list[str]]:
    if ncols < 6:
        return {}
    # column mapping (6열 / 10열 대응)
    if ncols >= 10:
        col_period, col_name, col_job, col_task, col_note = 0, 1, 6, 7, 9
        col_issuer, col_type, col_spec, col_pos = 1, 2, 6, 7
    else:
        col_period, col_name, col_note = 0, 1, ncols - 1
        col_job = max(3, ncols - 3)
        col_task = max(4, ncols - 2)
        col_issuer = 1
        col_type = 2 if ncols >= 3 else 1
        col_spec = col_job
        col_pos = col_task

    def row_with(pairs: list[tuple[int, str]]) -> list[str]:
        r = [""] * ncols
        for c, v in pairs:
            if 0 <= c < ncols:
                r[c] = v
        return r

    # Unicode escapes to avoid source/console encoding issues
    참여기간 = "\uCC38\uC5EC\uAE30\uAC04"
    인정일 = "\uC778\uC815\uC77C"
    참여일 = "\uCC38\uC5EC\uC77C"
    비고 = "\uBE44\uACE0"
    사업명 = "\uC0AC\uC5C5\uBA85"
    발주자 = "\uBC1C\uC8FC\uC790"
    공사종류 = "\uACF5\uC0AC\uC885\uB958"
    직무분야 = "\uC9C1\uBB34\uBD84\uC57C"
    전문분야 = "\uC804\uBB38\uBD84\uC57C"
    담당업무 = "\uB2F4\uB2F9\uC5C5\uBB34"
    책임정도 = "\uCC45\uC784\uC815\uB3C4"
    직위 = "\uC9C1\uC704"
    공사개요 = "\uACF5\uC0AC(\uC6A9\uC5ED)\uAC1C\uC694"
    공사금액 = "\uACF5\uC0AC(\uC6A9\uC5ED)\uAE08\uC561(\uBC31\uB9CC\uC6D0)"
    적용공법 = "\uC801\uC6A9 \uACF5\uBC95"
    적용융복합 = "\uC801\uC6A9 \uC735\u30FB\uBCF5\uD569\uAC74\uC124\uAE30\uC220"
    적용신기술 = "\uC801\uC6A9 \uC2E0\uAE30\uC220 \uB4F1"
    시설물종류 = "\uC2DC\uC124\uBB3C \uC885\uB958"

    return {
        0: row_with(
            [
                (col_period, f"{참여기간}\n({인정일})\n({참여일})"),
                (col_name, 사업명),
                (col_job, 직무분야),
                (col_task, 담당업무),
                (col_note, 비고),
            ]
        ),
        1: row_with([(col_issuer, 발주자), (col_type, 공사종류), (col_spec, 전문분야), (col_pos, 직위)]),
        2: row_with([(col_issuer, 공사개요), (col_spec, 책임정도), (col_pos, 공사금액)]),
        3: row_with([(col_issuer, 적용공법), (col_type, 적용융복합), (col_spec, 적용신기술), (col_pos, 시설물종류)]),
    }


def _write_blocks_on_sheet(
    ws, tables: list[list[list[object]]], prefix: str, *, clean: bool, recover_headers: bool
) -> None:
    row = 1
    for ti, table in enumerate(tables):
        if not table:
            continue
        t_use = clean_table(table) if clean else table
        nrows = len(t_use)
        ncols = max((len(r) for r in t_use if r), default=0)
        # NOTE: 문자열이 '='로 시작하면 Excel이 수식으로 오인할 수 있어, 구분 헤더는 '='로 시작하지 않게 한다.
        ws.cell(row=row, column=1, value=f"[TABLE] {prefix} #{ti}  ({nrows} x {ncols})")
        row += 1
        start = _infer_header_start_row(t_use) if recover_headers else None
        overrides = _header_override_rows(ncols) if (recover_headers and start is not None) else {}
        for tri, tr in enumerate(t_use):
            use = tr
            if start is not None:
                rel = tri - start
                if rel in overrides:
                    use = overrides[rel]
            for cj, cell in enumerate(use or []):
                ws.cell(row=row, column=cj + 1, value=_cell_str(cell))
            row += 1
        row += 2


def _autofit(ws, max_w: float = 55.0) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        m = 8
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                m = min(max(m, len(str(v)) + 1), int(max_w))
        ws.column_dimensions[letter].width = float(m)


def export_merged_by_page(
    pdf_path: Path,
    out_path: Path,
    page_indices: list[int],
    *,
    clean: bool,
    recover_headers: bool,
) -> Path:
    used: set[str] = set()
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    idx_ws = wb.create_sheet(_safe_sheet_name(used, "INDEX"), 0)
    idx_ws.append(["page_idx", "page_no", "merged_table_count", "note"])

    with pdfplumber.open(str(pdf_path)) as pdf:
        for pi in page_indices:
            page = pdf.pages[pi]
            tables_all = extract_tables_merged(page)
            tables = _pick_best_table_only(tables_all)
            idx_ws.append(
                [
                    pi,
                    pi + 1,
                    len(tables_all),
                    "best_only=1 (merged candidates scored; title/note tables penalized)",
                ]
            )
            sheet_name = _safe_sheet_name(used, f"p{pi + 1:03d}")
            ws = wb.create_sheet(sheet_name)
            ws.cell(row=1, column=1, value=f"PDF page index {pi} (human page {pi + 1}) — extract_tables_merged")
            _write_blocks_on_sheet(ws, tables, "merged", clean=clean, recover_headers=recover_headers)
            _autofit(ws)

    _autofit(idx_ws)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def export_by_strategy_by_page(
    pdf_path: Path,
    out_path: Path,
    page_indices: list[int],
    *,
    clean: bool,
    recover_headers: bool,
) -> Path:
    used: set[str] = set()
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    idx_ws = wb.create_sheet(_safe_sheet_name(used, "INDEX"), 0)
    idx_ws.append(["page_idx", "page_no", "n_line", "n_text", "n_default"])

    with pdfplumber.open(str(pdf_path)) as pdf:
        for pi in page_indices:
            page = pdf.pages[pi]
            t_line = safe_extract_tables(page, LINE_TABLE_SETTINGS)
            t_text = safe_extract_tables(page, TEXT_TABLE_SETTINGS)
            t_def = safe_extract_tables(page, None)
            idx_ws.append([pi, pi + 1, len(t_line), len(t_text), len(t_def)])

            sheet_name = _safe_sheet_name(used, f"p{pi + 1:03d}_st")
            ws = wb.create_sheet(sheet_name)
            r = 1
            ws.cell(row=r, column=1, value=f"page {pi + 1} — LINE (explicit_vertical_lines + lines)")
            r += 1
            for block in (
                (t_line, "LINE"),
                (t_text, "TEXT"),
                (t_def, "DEFAULT"),
            ):
                tbls, label = block
                ws.cell(row=r, column=1, value=f"--- {label} ({len(tbls)} tables) ---")
                r += 1
                for ti, table in enumerate(tbls):
                    if not table:
                        continue
                    t_use = clean_table(table) if clean else table
                    nrows = len(t_use)
                    ncols = max((len(x) for x in t_use if x), default=0)
                    ws.cell(row=r, column=1, value=f"[TABLE] {label} #{ti} ({nrows}x{ncols})")
                    r += 1
                    start = _infer_header_start_row(t_use) if recover_headers else None
                    overrides = _header_override_rows(ncols) if (recover_headers and start is not None) else {}
                    for tri, tr in enumerate(t_use):
                        use = tr
                        if start is not None:
                            rel = tri - start
                            if rel in overrides:
                                use = overrides[rel]
                        for cj, cell in enumerate(use or []):
                            ws.cell(row=r, column=cj + 1, value=_cell_str(cell))
                        r += 1
                    r += 1
                r += 1
            _autofit(ws)

    _autofit(idx_ws)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def run_export_job(
    pdf_path: Path,
    out_path: Path,
    pages_spec: str | None,
    by_strategy: bool,
    clean: bool = True,
    recover_headers: bool = False,
) -> tuple[Path | None, str]:
    """
    CLI/GUI 공통 실행 본문.
    Returns:
        (저장 경로, 오류 메시지) — 성공 시 오류 메시지는 "".
    """
    if not pdf_path.is_file():
        return None, f"파일 없음: {pdf_path}"
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            total = len(pdf.pages)
    except Exception as e:
        return None, f"PDF 열기 실패: {e}"
    pages = _parse_page_range(pages_spec, total)
    if not pages:
        return None, "유효한 페이지가 없습니다."
    try:
        if by_strategy:
            path = export_by_strategy_by_page(pdf_path, out_path, pages, clean=clean, recover_headers=recover_headers)
        else:
            path = export_merged_by_page(pdf_path, out_path, pages, clean=clean, recover_headers=recover_headers)
        return path, ""
    except Exception as e:
        return None, str(e)


def main() -> int:
    ap = argparse.ArgumentParser(description="pdfplumber 인식 표(파싱 전)를 Excel로 덤프")
    ap.add_argument("pdf", type=Path, help="PDF 경로")
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="출력 .xlsx (기본: PDF와 같은 폴더, 파일명_stem_recognized_tables.xlsx)",
    )
    ap.add_argument(
        "--pages",
        type=str,
        default=None,
        help="페이지 범위. 예: 0-2 또는 0 1 2 (0부터). 미지정 시 전체",
    )
    ap.add_argument(
        "--by-strategy",
        action="store_true",
        help="merged 대신 LINE / TEXT / DEFAULT 전략별 블록을 한 시트에 세로로 나열",
    )
    ap.add_argument(
        "--raw",
        action="store_true",
        help="정제(clean) 없이 원시 테이블을 그대로 출력합니다",
    )
    ap.add_argument(
        "--recover-headers",
        action="store_true",
        help="헤더 라벨(참여기간/비고 등)을 표 구조에 맞춰 정상 한글로 덮어써서 출력합니다",
    )
    args = ap.parse_args()

    pdf_path = args.pdf
    out = args.output
    if out is None:
        out = pdf_path.parent / f"{pdf_path.stem}_recognized_tables.xlsx"

    path, err = run_export_job(
        pdf_path,
        out,
        args.pages,
        args.by_strategy,
        clean=(not args.raw),
        recover_headers=bool(args.recover_headers),
    )
    if err or path is None:
        print(f"[ERROR] {err}", file=sys.stderr)
        return 2

    print(f"[OK] 저장: {path.resolve()}")
    with pdfplumber.open(str(pdf_path)) as pdf:
        total = len(pdf.pages)
    pages = _parse_page_range(args.pages, total)
    if pages:
        print(f"[INFO] 페이지 수: {len(pages)} (index: {pages[0]}..{pages[-1]})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
