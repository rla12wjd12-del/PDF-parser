#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
특정 좌/우 가상선(left/right)을 주고 lines 전략으로만 표를 추출해 Excel로 저장한다.
검색된 좌표(L/R)가 실제로 '참여기간/비고' 열을 분리하는지 시각적으로 확인하는 용도.

사용:
  python scripts/export_tables_with_custom_borders.py "originalPDF/황규철 경력증명서(2025.07.24).pdf" --left 27 --right 560
  python scripts/export_tables_with_custom_borders.py "..." --left 27 --right 560 --pages 4 10 -o excel_output/out.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from parsers.table_settings import LINE_TABLE_SETTINGS, safe_extract_tables  # noqa: E402


def _cell(v: object) -> str:
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").replace("\r", "\n")


def _autofit(ws, max_w: float = 55.0) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        m = 8
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                m = min(max(m, len(str(v)) + 1), int(max_w))
        ws.column_dimensions[letter].width = float(m)


def _try_load_fitz():
    try:
        import fitz  # type: ignore

        return fitz
    except Exception:
        return None


def _fitz_tokens_in_top_band(pdf_path: Path, page_idx: int, *, top: float = 260.0) -> set[str]:
    """
    PyMuPDF로 상단 밴드 텍스트를 읽어 토큰 집합을 만든다(문자 깨짐 회피 목적).
    """
    fitz = _try_load_fitz()
    if fitz is None:
        return set()
    doc = fitz.open(str(pdf_path))
    try:
        page = doc.load_page(page_idx)
        rect = fitz.Rect(0, 0, page.rect.width, min(top, page.rect.height))
        txt = page.get_text("text", clip=rect) or ""
        txt = re.sub(r"\s+", " ", txt)
        toks = set()
        for k in [
            "참여기간",
            "인정일",
            "참여일",
            "사업명",
            "발주자",
            "직무분야",
            "전문분야",
            "직위",
            "담당업무",
            "비고",
            "공사(용역)개요",
            "책임정도",
            "공사(용역)금액",
            "적용 공법",
            "적용 융",
            "적용 신기술",
            "시설물 종류",
        ]:
            if k in txt:
                toks.add(k)
        return toks
    finally:
        doc.close()


def _infer_header_start_row(table: list[list[object]]) -> int | None:
    if not table:
        return None
    date_like = re.compile(r"^\s*\d{4}\.\d{2}\.\d{2}\b")
    for i, r in enumerate(table[:12]):
        if i == 0:
            continue
        if not r:
            continue
        non_empty = [str(c).strip() for c in r if str(c or "").strip()]
        if len(non_empty) < 3:
            continue
        c0 = str(r[0] or "").strip() if len(r) >= 1 else ""
        if c0 and date_like.match(c0):
            continue
        return i
    return None


def _header_override_rows(max_cols: int) -> dict[int, list[str]]:
    """
    0-based(테이블 내부) 인덱스 기준: start, start+1, start+2, start+3 에 들어갈 헤더 텍스트.
    표 구조는 그대로 두고 '보이는 라벨'만 채운다.
    """
    if max_cols < 6:
        return {}
    # column mapping
    if max_cols >= 10:
        col_period, col_name, col_job, col_task, col_note = 0, 1, 6, 7, 9
        col_issuer, col_type, col_spec, col_pos = 1, 2, 6, 7
    else:
        col_period, col_name, col_note = 0, 1, max_cols - 1
        col_job = max(3, max_cols - 3)
        col_task = max(4, max_cols - 2)
        col_issuer = 1
        col_type = 2 if max_cols >= 3 else 1
        col_spec = col_job
        col_pos = col_task

    def row_with(pairs: list[tuple[int, str]]) -> list[str]:
        r = [""] * max_cols
        for c, v in pairs:
            if 0 <= c < max_cols:
                r[c] = v
        return r

    # Unicode escapes to avoid any source/console encoding issues
    참여기간 = "\uCC38\uC5EC\uAE30\uAC04"
    인정일 = "\uC778\uC815\uC77C"
    참여일 = "\uCC38\uC5EC\uC77C"
    비고 = "\uBE44\uACE0"
    사업명 = "\uC0AC\uC5C5\uBA85"
    발주자 = "\uBC1C\uC8FC\uC790"
    공사종류 = "\uACF5\uC0AC\uC885\uB958"
    직무분야 = "\uC9C1\uBB34\uBD84\uC57C"
    전문분야 = "\uC804\uBB38\uBD84\uC57C"
    담당업무 = "\uB2F4\uB2F9\uC5C5\uBB34"
    책임정도 = "\uCC45\uC784\uC815\uB3C4"
    직위 = "\uC9C1\uC704"
    공사개요 = "\uACF5\uC0AC(\uC6A9\uC5ED)\uAC1C\uC694"
    공사금액 = "\uACF5\uC0AC(\uC6A9\uC5ED)\uAE08\uC561(\uBC31\uB9CC\uC6D0)"
    적용공법 = "\uC801\uC6A9 \uACF5\uBC95"
    적용융복합 = "\uC801\uC6A9 \uC735\u30FB\uBCF5\uD569\uAC74\uC124\uAE30\uC220"
    적용신기술 = "\uC801\uC6A9 \uC2E0\uAE30\uC220 \uB4F1"
    시설물종류 = "\uC2DC\uC124\uBB3C \uC885\uB958"

    return {
        0: row_with(
            [
                (col_period, f"{참여기간}\n({인정일})\n({참여일})"),
                (col_name, 사업명),
                (col_job, 직무분야),
                (col_task, 담당업무),
                (col_note, 비고),
            ]
        ),
        1: row_with([(col_issuer, 발주자), (col_type, 공사종류), (col_spec, 전문분야), (col_pos, 직위)]),
        2: row_with([(col_issuer, 공사개요), (col_spec, 책임정도), (col_pos, 공사금액)]),
        3: row_with(
            [
                (col_issuer, 적용공법),
                (col_type, 적용융복합),
                (col_spec, 적용신기술),
                (col_pos, 시설물종류),
            ]
        ),
    }


def main() -> int:
    ap = argparse.ArgumentParser(description="custom borders(lines) → tables → xlsx")
    ap.add_argument("pdf", type=Path)
    ap.add_argument("--left", type=float, required=True)
    ap.add_argument("--right", type=float, required=True)
    ap.add_argument("--pages", nargs="*", type=int, default=None, help="0-based page indices (미지정 시 전체)")
    ap.add_argument("-o", "--output", type=Path, default=None)
    ap.add_argument(
        "--recover-headers",
        action="store_true",
        help="PyMuPDF로 헤더(참여기간/비고 등)를 복원해 테이블 헤더 셀에 덮어씁니다",
    )
    args = ap.parse_args()

    if not args.pdf.is_file():
        print(f"[ERROR] file not found: {args.pdf}", file=sys.stderr)
        return 2

    out = args.output
    if out is None:
        out = Path("excel_output") / f"{args.pdf.stem}_L{args.left:g}_R{args.right:g}_lines.xlsx"

    wb = Workbook()
    idx = wb.active
    idx.title = "INDEX"
    idx.append(["page_idx", "page_no", "table_count"])

    with pdfplumber.open(str(args.pdf)) as pdf:
        total = len(pdf.pages)
        pages = list(range(total)) if not args.pages else [p for p in args.pages if 0 <= p < total]
        for pi in pages:
            page = pdf.pages[pi]
            st = dict(LINE_TABLE_SETTINGS)
            st["explicit_vertical_lines"] = [float(args.left), float(args.right)]
            tabs = safe_extract_tables(page, st)
            idx.append([pi, pi + 1, len(tabs)])

            ws = wb.create_sheet(f"p{pi + 1:03d}")
            row = 1
            ws.cell(row=row, column=1, value=f"lines + explicit_vertical_lines=[{args.left}, {args.right}]")
            row += 2
            tokens = _fitz_tokens_in_top_band(args.pdf, pi) if args.recover_headers else set()
            for ti, t in enumerate(tabs):
                if not t:
                    continue
                max_cols = max((len(r) for r in t if r), default=0)
                ws.cell(row=row, column=1, value=f"[TABLE] {ti} ({len(t)} x {max_cols})")
                row += 1
                start = _infer_header_start_row(t) if args.recover_headers else None
                overrides = _header_override_rows(max_cols) if (args.recover_headers and start is not None) else {}
                for tri, tr in enumerate(t):
                    use = tr
                    if start is not None:
                        rel = tri - start
                        if rel in overrides:
                            use = overrides[rel]
                    for cj, c in enumerate(use or []):
                        ws.cell(row=row, column=cj + 1, value=_cell(c))
                    row += 1
                row += 2
            _autofit(ws)

    _autofit(idx)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"[OK] saved: {out.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

