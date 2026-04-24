#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
원본 page_2_parser 와 표 단독 page_2_parser_table_only 의 결과를 비교한다.

사용법
------
    # 단일 PDF 비교
    python compare_page_2_parsers.py "C:\\path\\to\\sample.pdf"

    # 폴더 내 모든 PDF 비교
    python compare_page_2_parsers.py "C:\\path\\to\\pdf_folder"

    # 출력 엑셀 파일명 지정
    python compare_page_2_parsers.py "C:\\path\\to\\sample.pdf" --out compare_result.xlsx

출력
----
    엑셀 파일 1개. 시트 구성:
    - "summary"            : 파일별 행 수 / 페이지별 행 수 비교
    - "row_diff"           : 두 파서 결과 행을 (참여기간_시작일, 참여기간_종료일) 키로 매칭한 차이
    - "원본_only"          : 원본에만 있는 행
    - "표단독_only"        : 표 단독에만 있는 행
    - "원본_full"          : 원본 결과 전체 덤프
    - "표단독_full"        : 표 단독 결과 전체 덤프

전제
----
- 본 스크립트는 프로젝트 루트(parsers/, data/ 등 폴더가 있는 위치)에서 실행한다.
- parsers/page_2_parser.py 와 parsers/page_2_parser_table_only.py 두 개가 모두 존재해야 한다.
- 기존 프로젝트의 DocumentContext 를 그대로 사용한다 (PDF 로딩 방식 일치).
"""

import argparse
import os
import sys
from pathlib import Path
from typing import List, Dict, Any, Tuple

# 프로젝트 루트를 sys.path에 등록 (compare_*.py 가 루트에 있다는 가정)
_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(_ROOT))

try:
    from parsers.document_context import DocumentContext
except Exception as e:
    print(f"[FATAL] DocumentContext import 실패: {e}")
    print("       이 스크립트는 프로젝트 루트(parsers/ 폴더가 있는 위치)에서 실행해야 합니다.")
    sys.exit(1)

# 두 파서 import
from parsers.page_2_parser import parse_page_2 as parse_page_2_original
from parsers.page_2_parser_table_only import parse_page_2 as parse_page_2_table_only

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("[FATAL] openpyxl 이 설치되어 있지 않습니다. `pip install openpyxl` 후 재실행해주세요.")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# 비교 대상 필드 (사업명/발주자/직위 등 핵심 18종)
# ─────────────────────────────────────────────────────────────────────────────
COMPARE_FIELDS = [
    "참여기간_시작일",
    "참여기간_종료일",
    "인정일수",
    "참여일수",
    "사업명",
    "발주자",
    "공사종류",
    "직무분야",
    "전문분야",
    "담당업무",
    "책임정도",
    "직위",
    "공사(용역)금액(백만원)",
    "공사(용역)개요",
    "적용 공법",
    "적용 융복합건설기술",
    "적용 신기술 등",
    "시설물 종류",
    "비고",
]


def _norm(v: Any) -> str:
    """비교용 정규화: None/공백 → 빈문자열, 줄바꿈/연속공백 정리"""
    if v is None:
        return ""
    s = str(v)
    s = s.replace("\r", " ").replace("\n", " ")
    while "  " in s:
        s = s.replace("  ", " ")
    return s.strip()


def _row_key(r: Dict[str, Any]) -> Tuple[str, str]:
    """행 매칭 키: (시작일, 종료일). 사업명은 두 파서 간 다를 수 있어 키로 안 씀."""
    return (_norm(r.get("참여기간_시작일")), _norm(r.get("참여기간_종료일")))


def _parse_one_pdf(pdf_path: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[int, Tuple[int, int]]]:
    """
    단일 PDF에 두 파서를 적용해 (원본_rows, 표단독_rows, 페이지별_행수_dict) 반환.
    페이지별_행수_dict[페이지(1-based)] = (원본_n, 표단독_n)
    """
    rows_orig: List[Dict[str, Any]] = []
    rows_table: List[Dict[str, Any]] = []
    page_counts: Dict[int, Tuple[int, int]] = {}

    try:
        with DocumentContext.open(pdf_path) as ctx_orig, DocumentContext.open(pdf_path) as ctx_table:
            total_pages = ctx_orig.total_pages
            # 기술경력은 일반적으로 2페이지부터 시작 (page_num=1, 1-based=2)
            # 안전하게 페이지 0~끝까지 모두 시도하되, 실제 데이터가 나온 페이지만 기록
            for page_num in range(total_pages):
                try:
                    o = parse_page_2_original(ctx_orig, page_num) or []
                except Exception as e:
                    print(f"  [WARN] 원본 파서 페이지 {page_num + 1} 예외: {e}")
                    o = []
                try:
                    t = parse_page_2_table_only(ctx_table, page_num) or []
                except Exception as e:
                    print(f"  [WARN] 표단독 파서 페이지 {page_num + 1} 예외: {e}")
                    t = []

                if o or t:
                    page_counts[page_num + 1] = (len(o), len(t))
                    for r in o:
                        if isinstance(r, dict):
                            r.setdefault("_pdf_pages", [page_num + 1])
                            r["_pdf_file"] = os.path.basename(pdf_path)
                            rows_orig.append(r)
                    for r in t:
                        if isinstance(r, dict):
                            r.setdefault("_pdf_pages", [page_num + 1])
                            r["_pdf_file"] = os.path.basename(pdf_path)
                            rows_table.append(r)
    except Exception as e:
        print(f"  [ERROR] PDF 처리 중 오류 발생: {e}")
        return rows_orig, rows_table, page_counts

    return rows_orig, rows_table, page_counts


def _diff_rows(rows_orig: List[Dict[str, Any]], rows_table: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    두 결과를 (시작일, 종료일) 키로 매칭해 차이를 계산.
    Returns:
        {
            "matched_diffs": [{"key": (s,e), "field": ..., "원본": ..., "표단독": ..., ...}, ...],
            "only_in_orig":  [원본에만 있는 행, ...],
            "only_in_table": [표단독에만 있는 행, ...],
        }
    """
    by_key_orig: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    by_key_table: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}

    for r in rows_orig:
        by_key_orig.setdefault(_row_key(r), []).append(r)
    for r in rows_table:
        by_key_table.setdefault(_row_key(r), []).append(r)

    matched_diffs: List[Dict[str, Any]] = []
    only_in_orig: List[Dict[str, Any]] = []
    only_in_table: List[Dict[str, Any]] = []

    all_keys = set(by_key_orig.keys()) | set(by_key_table.keys())
    for k in sorted(all_keys):
        os_ = by_key_orig.get(k, [])
        ts_ = by_key_table.get(k, [])
        # 동일 키에 여러 행이 있을 수 있음(매우 드물지만 동일 시작/종료일 다른 사업)
        # 단순화를 위해 인덱스별로 페어링
        n = max(len(os_), len(ts_))
        for i in range(n):
            o = os_[i] if i < len(os_) else None
            t = ts_[i] if i < len(ts_) else None
            if o is None and t is not None:
                only_in_table.append(t)
                continue
            if t is None and o is None:
                continue
            if t is None and o is not None:
                only_in_orig.append(o)
                continue
            # 양쪽 다 존재 → 필드별 비교
            for f in COMPARE_FIELDS:
                vo = _norm(o.get(f))
                vt = _norm(t.get(f))
                if vo != vt:
                    matched_diffs.append({
                        "_pdf_file": o.get("_pdf_file") or t.get("_pdf_file") or "",
                        "_pdf_pages": str(o.get("_pdf_pages") or t.get("_pdf_pages") or ""),
                        "참여기간_시작일": k[0],
                        "참여기간_종료일": k[1],
                        "field": f,
                        "원본": vo,
                        "표단독": vt,
                    })

    return {
        "matched_diffs": matched_diffs,
        "only_in_orig": only_in_orig,
        "only_in_table": only_in_table,
    }


def _write_excel(
    out_path: str,
    summary_rows: List[Dict[str, Any]],
    diff: Dict[str, Any],
    rows_orig: List[Dict[str, Any]],
    rows_table: List[Dict[str, Any]],
) -> None:
    wb = openpyxl.Workbook()

    # ---- summary ----
    ws = wb.active
    ws.title = "summary"
    headers = ["pdf_file", "page", "원본_n", "표단독_n", "차이"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    for i, r in enumerate(summary_rows, 2):
        ws.cell(row=i, column=1, value=r["pdf_file"])
        ws.cell(row=i, column=2, value=r["page"])
        ws.cell(row=i, column=3, value=r["원본_n"])
        ws.cell(row=i, column=4, value=r["표단독_n"])
        diff_val = r["원본_n"] - r["표단독_n"]
        c = ws.cell(row=i, column=5, value=diff_val)
        if diff_val != 0:
            c.fill = PatternFill("solid", fgColor="FFE699")
    for col_letter, width in zip("ABCDE", [40, 8, 10, 10, 8]):
        ws.column_dimensions[col_letter].width = width

    # ---- row_diff ----
    ws = wb.create_sheet("row_diff")
    headers = ["pdf_file", "page", "참여기간_시작일", "참여기간_종료일", "field", "원본", "표단독"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    for i, r in enumerate(diff["matched_diffs"], 2):
        ws.cell(row=i, column=1, value=r["_pdf_file"])
        ws.cell(row=i, column=2, value=r["_pdf_pages"])
        ws.cell(row=i, column=3, value=r["참여기간_시작일"])
        ws.cell(row=i, column=4, value=r["참여기간_종료일"])
        ws.cell(row=i, column=5, value=r["field"])
        ws.cell(row=i, column=6, value=r["원본"])
        ws.cell(row=i, column=7, value=r["표단독"])
    for col_letter, width in zip("ABCDEFG", [30, 10, 14, 14, 22, 50, 50]):
        ws.column_dimensions[col_letter].width = width

    # ---- only_in_orig / only_in_table ----
    def _dump_rows_sheet(title: str, rows: List[Dict[str, Any]]):
        ws2 = wb.create_sheet(title)
        cols = ["_pdf_file", "_pdf_pages"] + COMPARE_FIELDS
        for j, h in enumerate(cols, 1):
            c = ws2.cell(row=1, column=j, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="FFD7D7" if "orig" in title else "D7F0DD")
        for i, r in enumerate(rows, 2):
            for j, h in enumerate(cols, 1):
                v = r.get(h, "")
                if h == "_pdf_pages":
                    v = str(v)
                ws2.cell(row=i, column=j, value=_norm(v) if h not in ("_pdf_file", "_pdf_pages") else v)

    _dump_rows_sheet("원본_only", diff["only_in_orig"])
    _dump_rows_sheet("표단독_only", diff["only_in_table"])

    # ---- 전체 덤프 ----
    def _dump_full_sheet(title: str, rows: List[Dict[str, Any]]):
        ws2 = wb.create_sheet(title)
        cols = ["_pdf_file", "_pdf_pages"] + COMPARE_FIELDS
        for j, h in enumerate(cols, 1):
            c = ws2.cell(row=1, column=j, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="D9E1F2")
        for i, r in enumerate(rows, 2):
            for j, h in enumerate(cols, 1):
                v = r.get(h, "")
                if h == "_pdf_pages":
                    v = str(v)
                ws2.cell(row=i, column=j, value=_norm(v) if h not in ("_pdf_file", "_pdf_pages") else v)

    _dump_full_sheet("원본_full", rows_orig)
    _dump_full_sheet("표단독_full", rows_table)

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description="원본/표단독 page_2 파서 결과 비교")
    ap.add_argument("input", help="비교할 PDF 파일 또는 폴더 경로")
    ap.add_argument("--out", default="compare_page_2_result.xlsx", help="출력 엑셀 파일명")
    args = ap.parse_args()

    in_path = Path(args.input)
    pdfs: List[Path] = []
    if in_path.is_dir():
        pdfs = sorted(in_path.rglob("*.pdf"))
    elif in_path.is_file() and in_path.suffix.lower() == ".pdf":
        pdfs = [in_path]
    else:
        print(f"[FATAL] 입력 경로가 PDF 파일 또는 폴더가 아닙니다: {in_path}")
        sys.exit(1)

    if not pdfs:
        print(f"[INFO] PDF 파일을 찾지 못했습니다: {in_path}")
        sys.exit(0)

    all_rows_orig: List[Dict[str, Any]] = []
    all_rows_table: List[Dict[str, Any]] = []
    summary_rows: List[Dict[str, Any]] = []

    for pdf in pdfs:
        print(f"\n=== {pdf.name} ===")
        rows_o, rows_t, page_counts = _parse_one_pdf(str(pdf))
        all_rows_orig.extend(rows_o)
        all_rows_table.extend(rows_t)
        for page, (n_o, n_t) in sorted(page_counts.items()):
            summary_rows.append({
                "pdf_file": pdf.name,
                "page": page,
                "원본_n": n_o,
                "표단독_n": n_t,
            })
        print(f"  → 원본 총 {len(rows_o)}행 / 표단독 총 {len(rows_t)}행")

    diff = _diff_rows(all_rows_orig, all_rows_table)

    print(f"\n[결과 요약]")
    print(f"  matched 행 차이(필드 단위): {len(diff['matched_diffs'])}건")
    print(f"  원본에만 있는 행          : {len(diff['only_in_orig'])}건")
    print(f"  표단독에만 있는 행        : {len(diff['only_in_table'])}건")

    out_path = str(Path(args.out).resolve())
    _write_excel(out_path, summary_rows, diff, all_rows_orig, all_rows_table)
    print(f"\n[OK] 비교 엑셀 저장: {out_path}")


if __name__ == "__main__":
    main()
