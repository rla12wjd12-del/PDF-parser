from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl
import json


_GRADE_TOKENS = ("특급", "고급", "중급", "초급")


def _norm(s: str) -> str:
    """
    PDF/엑셀/텍스트 추출에서 흔히 섞이는 구분점(·/ㆍ)과 공백을 통일한다.
    - 'ㆍ' (U+318D) → '·' (U+00B7)
    - 연속 공백 축약, 양끝 공백 제거
    """
    if s is None:
        return ""
    s = str(s)
    s = s.replace("ㆍ", "·")
    s = " ".join(s.split())
    return s.strip()


@dataclass(frozen=True)
class FieldCatalog:
    job_fields: Tuple[str, ...]
    specialty_by_job: Dict[str, Tuple[str, ...]]

    @property
    def all_specialties(self) -> Tuple[str, ...]:
        seen = []
        seen_set = set()
        for job in self.job_fields:
            for sp in self.specialty_by_job.get(job, ()):
                if sp not in seen_set:
                    seen.append(sp)
                    seen_set.add(sp)
        return tuple(seen)


def _default_catalog() -> FieldCatalog:
    # 엑셀을 못 읽는 경우를 대비한 최소 폴백(기존 상수 기반 + '·' 표기 통일)
    specialty_by_job = {
        "토목": (
            "토질·지질",
            "토목구조",
            "항만 및 해안",
            "도로 및 공항",
            "철도·삭도",
            "수자원개발",
            "상하수도",
            "농어업토목",
            "토목시공",
            "토목품질관리",
            "측량 및 지형공간정보",
            "지적",
        ),
        "건축": (
            "건축구조",
            "건축기계설비",
            "건축시공",
            "실내건축",
            "건축품질관리",
            "건축계획·설계",
        ),
        "조경": ("조경계획", "조경시공관리"),
    }
    job_fields = (
        "토목",
        "건축",
        "조경",
        "기계",
        "전기",
        "환경",
        "안전",
        "정보통신",
        "화공",
    )
    # 표기 통일
    specialty_by_job = {k: tuple(_norm(v) for v in vs) for k, vs in specialty_by_job.items()}
    return FieldCatalog(job_fields=tuple(job_fields), specialty_by_job=specialty_by_job)


def _pick_catalog_xlsx(project_root: Path) -> Optional[Path]:
    """
    프로젝트 루트에서 '직무/전문' 관련 xlsx를 우선 탐색하고,
    없으면 첫 번째 xlsx를 사용한다.
    """
    if not project_root.exists():
        return None

    xlsx_files = list(project_root.glob("*.xlsx"))
    if not xlsx_files:
        return None

    # 파일명이 깨져 보이는 환경도 있어(콘솔 코드페이지) 내용 기반 검색은 어렵고,
    # 가능한 경우에만 이름 힌트를 사용한다.
    preferred = []
    for p in xlsx_files:
        name = p.name
        if ("직무" in name) or ("전문" in name) or ("분야" in name):
            preferred.append(p)
    return preferred[0] if preferred else xlsx_files[0]


def _pick_catalog_json(project_root: Path) -> Optional[Path]:
    """
    사람이 읽기 쉬운 형태(예: tools/build_field_catalog.py로 생성)인
    data/field_catalog.json이 있으면 우선 사용한다.
    """
    p = project_root / "data" / "field_catalog.json"
    return p if p.exists() else None


def _split_specialties(cell_value: str) -> List[str]:
    """
    엑셀 셀에 'A, B, C' 형태로 들어있는 전문분야를 리스트로 분리.
    """
    s = _norm(cell_value)
    if not s:
        return []
    # 엑셀에 콤마/줄바꿈 혼재 가능
    raw_parts: List[str] = []
    for chunk in s.replace("\n", ",").split(","):
        chunk = _norm(chunk)
        if chunk:
            raw_parts.append(chunk)
    return raw_parts


def _load_catalog_from_xlsx(xlsx_path: Path) -> FieldCatalog:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # 기대 포맷(현재 제공 파일): A열=직무분야, B열=전문분야(콤마로 구분)
    # 1행은 헤더인 경우가 많으므로 헤더 여부를 유연하게 판단한다.
    rows = []
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a is None and b is None:
            continue
        rows.append((_norm(a), _norm(b)))

    if not rows:
        return _default_catalog()

    start_idx = 0
    header_a, header_b = rows[0]
    if ("직무" in header_a) or ("전문" in header_b) or ("전문" in header_a):
        start_idx = 1

    job_fields: List[str] = []
    specialty_by_job: Dict[str, List[str]] = {}

    for job, specs in rows[start_idx:]:
        job = _norm(job)
        if not job:
            continue
        if job not in job_fields:
            job_fields.append(job)
        spec_list = _split_specialties(specs)
        if spec_list:
            specialty_by_job.setdefault(job, [])
            for sp in spec_list:
                sp = _norm(sp)
                if sp and sp not in specialty_by_job[job]:
                    specialty_by_job[job].append(sp)

    # tuple 고정 + 표기 통일
    specialty_by_job_t = {k: tuple(_norm(v) for v in vs) for k, vs in specialty_by_job.items()}
    return FieldCatalog(job_fields=tuple(job_fields), specialty_by_job=specialty_by_job_t)


def _load_catalog_from_json(json_path: Path) -> FieldCatalog:
    raw = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        return _default_catalog()
    job_fields = []
    specialty_by_job: Dict[str, Tuple[str, ...]] = {}
    for job, specs in raw.items():
        job = _norm(job)
        if not job:
            continue
        if job not in job_fields:
            job_fields.append(job)
        if isinstance(specs, list):
            specialty_by_job[job] = tuple(_norm(s) for s in specs if _norm(s))
    return FieldCatalog(job_fields=tuple(job_fields), specialty_by_job=specialty_by_job)


@lru_cache(maxsize=1)
def get_field_catalog(project_root: Optional[str] = None, xlsx_path: Optional[str] = None) -> FieldCatalog:
    root = Path(project_root) if project_root else Path(__file__).resolve().parent
    try:
        json_path = _pick_catalog_json(root)
        if json_path:
            return _load_catalog_from_json(json_path)

        xlsx = Path(xlsx_path) if xlsx_path else _pick_catalog_xlsx(root)
        if not xlsx or not xlsx.exists():
            return _default_catalog()
        return _load_catalog_from_xlsx(xlsx)
    except Exception:
        # 엑셀 포맷/권한/손상 등의 이유로 실패해도 파서가 죽지 않게 폴백
        return _default_catalog()


def best_match_specialty(text: str, catalog: Optional[FieldCatalog] = None) -> str:
    """
    주어진 텍스트(셀)에서 전문분야명을 '목록 기반'으로 가장 길게 매칭한다.
    - '토질·지질 특급' → '토질·지질'
    - 문서에 따라 'ㆍ'로 찍혀도 '·'로 통일해서 매칭한다.
    """
    t = _norm(text)
    if not t:
        return ""
    catalog = catalog or get_field_catalog()
    candidates = sorted((_norm(s) for s in catalog.all_specialties), key=len, reverse=True)
    for c in candidates:
        if c and c in t:
            return c
    return ""


def extract_name_and_grade(cell_text: str) -> Tuple[str, str]:
    """
    '... 특급/고급/중급/초급' 형태에서 등급 토큰을 찾아 (이름, 등급) 반환.
    이름은 원문에서 등급 토큰 앞부분을 반환(표기 통일 적용).
    """
    t = _norm(cell_text)
    if not t:
        return "", ""
    for g in _GRADE_TOKENS:
        if f" {g}" in t:
            name = _norm(t.split(f" {g}", 1)[0])
            return name, g
    return "", ""

