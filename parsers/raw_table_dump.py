#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
page2/page3 파서에서 사용되는 '원시 표'를 디버그용으로 엑셀로 덤프한다.

요구사항:
- 특정 PDF에만 맞춘 하드코딩 금지
- 파싱 중간 산출물을 excel_output 아래 새 폴더에 제공
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
import re

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# 한 번 실행(run) 동안 동일 폴더에 모으기 위한 run id
_RUN_ID = datetime.now().strftime("%Y%m%d_%H%M%S")


def _escape_excel_formula_text(v: Any) -> Any:
    if not isinstance(v, str):
        return v
    if not v:
        return v
    if v[0] in ("=", "+", "-", "@"):
        return "'" + v
    return v


def _safe_filename(name: str) -> str:
    s = str(name or "").strip()
    if not s:
        return "output"
    s = re.sub(r'[<>:"/\\\\|?*\\s]+', "_", s).strip("_")
    return s or "output"


def _autofit_columns(ws, max_width: float = 60.0) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), max_width)


def _truncate_cell(s: str, *, limit: int = 32000) -> str:
    # Excel 셀 최대 길이(대략 32767). 약간 보수적으로 제한.
    t = str(s or "")
    if len(t) <= limit:
        return t
    return t[: limit - 30] + " ...[TRUNCATED]..."


def _to_str_table(table: Any) -> list[list[str]]:
    if not table:
        return []
    out: list[list[str]] = []
    for r in (table or []):
        row = []
        for c in (r or []):
            v = "" if c is None else str(c)
            v = v.replace("\r\n", "\n").replace("\r", "\n")
            v = _truncate_cell(v)
            row.append(v)
        out.append(row)
    return out


def _write_table_sheet(ws, table: list[list[str]]) -> None:
    if not table:
        ws.append(["(표 없음)"])
        return
    max_cols = max((len(r) for r in table), default=0)
    if max_cols <= 0:
        ws.append(["(표 없음)"])
        return
    for r in table:
        rr = list(r) + [""] * (max_cols - len(r))
        ws.append([_escape_excel_formula_text(x) for x in rr])
    _autofit_columns(ws)


def _write_kv_sheet(ws, data: dict[str, Any]) -> None:
    ws.append(["key", "value"])
    for k, v in (data or {}).items():
        vv = v
        if isinstance(v, (list, dict)):
            vv = str(v)
        ws.append([_escape_excel_formula_text(str(k)), _escape_excel_formula_text(_truncate_cell(str(vv)))])
    _autofit_columns(ws)


def dump_raw_tables_to_excel(
    *,
    pdf_path: str,
    section: str,
    page_num_1based: int,
    tables_all: list[Any] | None,
    best_table: Any | None,
    normalized_6cols: list[list[str]] | None,
    meta: dict[str, Any] | None = None,
) -> Path | None:
    """
    원시표/정규화표/메타를 xlsx로 저장한다.

    저장 경로:
      <repo_root>/excel_output/raw_tables/<run_id>/<pdf_stem>/<section>_pNNN.xlsx
    """
    try:
        base_dir = Path(__file__).resolve().parents[1]
        out_root = base_dir / "excel_output" / "raw_tables" / f"raw_tables_{_RUN_ID}"

        pdf_stem = _safe_filename(Path(pdf_path).stem) if pdf_path else "unknown_pdf"
        out_dir = out_root / pdf_stem
        out_dir.mkdir(parents=True, exist_ok=True)

        sec = _safe_filename(section or "section")
        fname = f"{sec}_p{int(page_num_1based):03d}.xlsx"
        out_path = out_dir / fname

        wb = Workbook()
        ws0 = wb.active
        ws0.title = "meta"
        m = dict(meta or {})
        m.update(
            {
                "pdf_path": pdf_path or "",
                "pdf_stem": pdf_stem,
                "section": section or "",
                "page_num_1based": int(page_num_1based),
                "n_tables_all": int(len(tables_all or [])),
            }
        )
        _write_kv_sheet(ws0, m)

        ws1 = wb.create_sheet(title="raw_best")
        _write_table_sheet(ws1, _to_str_table(best_table))

        ws2 = wb.create_sheet(title="normalized_6cols")
        _write_table_sheet(ws2, _to_str_table(normalized_6cols))

        # 참고용: 전체 테이블 중 "상위 2개"만 추가(너무 커지는 것 방지)
        all_tabs = tables_all or []
        for i, t in enumerate(all_tabs[:2]):
            ws = wb.create_sheet(title=f"raw_all_{i+1}")
            _write_table_sheet(ws, _to_str_table(t))

        wb.save(out_path)
        return out_path.resolve()
    except Exception:
        return None

