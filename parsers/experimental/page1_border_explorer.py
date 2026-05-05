#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Page 1 가상 수직선 탐색기 ? pdfplumber 표 추출 시각화 도구.

전체 PDF에서 Page 1(인적사항 템플릿) 해당 페이지를 자동 감지한 뒤,
좌·우 가상 수직선 X 좌표를 슬라이더/드래그로 조절하면서
pdfplumber 표 추출 결과가 어떻게 달라지는지 실시간으로 확인한다.

기능:
  - 좌·우 가상 선 드래그(캔버스) + 슬라이더로 위치 조절
  - 추가 중간 선(콤마 구분 숫자 입력) 지원
  - 표 추출 전략 선택(lines / text / default)
  - PDF 원래 선 표시 / 표·셀 bbox 표시 토글
  - PDF 선 X 좌표 목록에 스냅(좌·우 각각)
  - 추출 결과 텍스트 패널에 열 맞춤 표 형식으로 표시
  - 현재 페이지 / 전체 페이지 Excel 저장

실행 (PDF-parser 루트):
  python parsers/experimental/page1_border_explorer.py
  python parsers/experimental/page1_border_explorer.py "path/to/file.pdf"
"""

from __future__ import annotations

import sys
import re
from pathlib import Path
from typing import Any, List, Optional

# ── 경로 설정 ────────────────────────────────────────────────────────────────
_HERE = Path(__file__).resolve().parent
_ROOT = _HERE.parent.parent  # PDF-parser root
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText

try:
    import fitz  # PyMuPDF
    _FITZ_OK = True
except ImportError:
    _FITZ_OK = False

try:
    from PIL import Image, ImageTk
    _PIL_OK = True
except ImportError:
    _PIL_OK = False

try:
    from openpyxl import Workbook
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

import pdfplumber

# ── 기본 가상선 좌표 ─────────────────────────────────────────────────────────
try:
    from parsers.utils.table_settings import VIRTUAL_LEFT_X, VIRTUAL_RIGHT_X
except Exception:
    VIRTUAL_LEFT_X: float = 27.0
    VIRTUAL_RIGHT_X: float = 560.0

# ── Page 1 판별 키워드 ────────────────────────────────────────────────────────
_PAGE1_KEYWORDS = ["인적사항", "성명(한글)", "국가기술자격", "교육훈련", "상훈"]

# ── 색상 ─────────────────────────────────────────────────────────────────────
COL_LEFT   = "#0055ff"   # 파란색 ? 좌선
COL_RIGHT  = "#cc0000"   # 빨간색 ? 우선
COL_MID    = "#009900"   # 초록색 ? 중간 추가선
COL_EDGE   = "#aaaaff"   # 연보라 ? PDF 원래 선
COL_TABLE  = "#ff8800"   # 주황  ? 표 bbox
COL_CELL   = "#ff8800"   # 주황  ? 셀 bbox

DRAG_THRESHOLD_PX = 10   # 드래그 감지 범위(픽셀)


class App:
    CANVAS_MIN_W = 580
    CANVAS_MIN_H = 820

    def __init__(self, root: tk.Tk, initial_pdf: Optional[str] = None) -> None:
        self.root = root
        self.root.title("Page 1 가상 수직선 탐색기")
        self.root.minsize(1100, 700)

        # ── 문서 상태
        self.pdf_path: Optional[Path] = None
        self.fitz_doc: Optional[Any] = None
        self.plumber_pdf: Optional[Any] = None
        self.page_indices: List[int] = []     # page-1 해당 PDF 인덱스 목록
        self.cur_pos = 0                       # page_indices 내 현재 위치

        # ── 렌더 상태
        self.scale: float = 1.0
        self.page_w_pdf: float = 595.0
        self.page_h_pdf: float = 842.0
        self._photo: Optional[Any] = None
        self._img_ox = 4   # 캔버스 내 이미지 x 오프셋(px)
        self._img_oy = 4   # 캔버스 내 이미지 y 오프셋(px)

        # ── PDF 수직선 X 목록 (스냅용)
        self._pdf_vline_xs: List[float] = []

        # ── 드래그 상태
        self._drag_target: Optional[str] = None   # 'left' | 'right'

        self._build_ui()

        if initial_pdf and Path(initial_pdf).is_file():
            self.pdf_path_var.set(str(initial_pdf))
            self._load_pdf()

    def _build_ui(self) -> None:
        root = self.root
        root.columnconfigure(0, weight=1)
        root.rowconfigure(1, weight=1)

        toolbar = ttk.Frame(root, padding=(6, 4))
        toolbar.grid(row=0, column=0, sticky=tk.EW)
        toolbar.columnconfigure(1, weight=1)

        ttk.Label(toolbar, text="PDF:").grid(row=0, column=0, sticky=tk.W)
        self.pdf_path_var = tk.StringVar()
        ttk.Entry(toolbar, textvariable=self.pdf_path_var, width=70).grid(
            row=0, column=1, sticky=tk.EW, padx=4
        )
        ttk.Button(toolbar, text="찾아보기", command=self._browse_pdf).grid(row=0, column=2)
        ttk.Button(toolbar, text="불러오기", command=self._load_pdf, style="Accent.TButton"
                   ).grid(row=0, column=3, padx=(4, 0))
        ttk.Button(toolbar, text="Excel 저장", command=self._save_current_page_to_excel
                   ).grid(row=0, column=4, padx=(4, 0))
        ttk.Button(toolbar, text="전체 Excel 저장", command=self._save_all_pages_to_excel
                   ).grid(row=0, column=5, padx=(4, 0))

        nav = ttk.Frame(toolbar)
        nav.grid(row=1, column=0, columnspan=6, sticky=tk.W, pady=3)

        ttk.Button(nav, text="◀ 이전", command=self._prev_page).pack(side=tk.LEFT, padx=2)
        self.page_label_var = tk.StringVar(value="페이지: -/-")
        ttk.Label(nav, textvariable=self.page_label_var, width=18).pack(side=tk.LEFT, padx=4)
        ttk.Button(nav, text="다음 ▶", command=self._next_page).pack(side=tk.LEFT, padx=2)

        ttk.Separator(nav, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)

        ttk.Label(nav, text="전략:").pack(side=tk.LEFT)
        self.strategy_var = tk.StringVar(value="lines")
        strat_cb = ttk.Combobox(
            nav, textvariable=self.strategy_var,
            values=["lines", "text", "default"],
            width=8, state="readonly",
        )
        strat_cb.pack(side=tk.LEFT, padx=4)
        strat_cb.bind("<<ComboboxSelected>>", lambda _: self._extract_and_update())

        ttk.Separator(nav, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)

        self.show_pdf_lines_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            nav, text="PDF 원래 선", variable=self.show_pdf_lines_var,
            command=self._redraw_overlays
        ).pack(side=tk.LEFT, padx=3)

        self.show_cells_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            nav, text="셀 bbox", variable=self.show_cells_var,
            command=self._extract_and_update
        ).pack(side=tk.LEFT, padx=3)

        sf = ttk.Frame(toolbar)
        sf.grid(row=2, column=0, columnspan=6, sticky=tk.EW, pady=2)
        sf.columnconfigure(2, weight=1)
        sf.columnconfigure(5, weight=1)

        ttk.Label(sf, text="좌선 X:", foreground=COL_LEFT, font=("", 9, "bold")).grid(
            row=0, column=0, sticky=tk.W, padx=(0, 4)
        )
        self.left_x_var = tk.DoubleVar(value=VIRTUAL_LEFT_X)
        self.left_slider = ttk.Scale(
            sf, from_=0, to=600, orient=tk.HORIZONTAL,
            variable=self.left_x_var, command=self._on_slider_change,
        )
        self.left_slider.grid(row=0, column=1, columnspan=2, sticky=tk.EW)
        self.left_x_entry = ttk.Entry(sf, width=7)
        self.left_x_entry.insert(0, f"{VIRTUAL_LEFT_X:.1f}")
        self.left_x_entry.grid(row=0, column=3, padx=4)
        self.left_x_entry.bind("<Return>", self._on_left_entry)
        self.left_x_entry.bind("<FocusOut>", self._on_left_entry)
        ttk.Button(sf, text="PDF에 스냅", command=lambda: self._snap_to_pdf_line("left"),
                   width=10).grid(row=0, column=4, padx=(0, 8))

        ttk.Label(sf, text="우선 X:", foreground=COL_RIGHT, font=("", 9, "bold")).grid(
            row=0, column=5, sticky=tk.W, padx=(8, 4)
        )
        self.right_x_var = tk.DoubleVar(value=VIRTUAL_RIGHT_X)
        self.right_slider = ttk.Scale(
            sf, from_=0, to=600, orient=tk.HORIZONTAL,
            variable=self.right_x_var, command=self._on_slider_change,
        )
        self.right_slider.grid(row=0, column=6, columnspan=2, sticky=tk.EW)
        sf.columnconfigure(7, weight=1)
        self.right_x_entry = ttk.Entry(sf, width=7)
        self.right_x_entry.insert(0, f"{VIRTUAL_RIGHT_X:.1f}")
        self.right_x_entry.grid(row=0, column=8, padx=4)
        self.right_x_entry.bind("<Return>", self._on_right_entry)
        self.right_x_entry.bind("<FocusOut>", self._on_right_entry)
        ttk.Button(sf, text="PDF에 스냅", command=lambda: self._snap_to_pdf_line("right"),
                   width=10).grid(row=0, column=9)

        af = ttk.Frame(toolbar)
        af.grid(row=3, column=0, columnspan=6, sticky=tk.EW, pady=2)
        ttk.Label(af, text="추가 선 X (콤마 구분):", foreground=COL_MID).pack(side=tk.LEFT)
        self.extra_lines_var = tk.StringVar(value="")
        extra_entry = ttk.Entry(af, textvariable=self.extra_lines_var, width=40)
        extra_entry.pack(side=tk.LEFT, padx=4)
        extra_entry.bind("<Return>", lambda _: self._extract_and_update())
        extra_entry.bind("<FocusOut>", lambda _: self._extract_and_update())
        ttk.Label(af, text="예: 150, 300, 450", foreground="gray").pack(side=tk.LEFT)

        paned = ttk.PanedWindow(root, orient=tk.HORIZONTAL)
        paned.grid(row=1, column=0, sticky=tk.NSEW, padx=4, pady=4)

        canvas_outer = ttk.LabelFrame(paned, text="PDF 페이지 뷰", padding=2)
        paned.add(canvas_outer, weight=1)

        self.canvas = tk.Canvas(
            canvas_outer,
            width=self.CANVAS_MIN_W, height=self.CANVAS_MIN_H,
            bg="#d8d8d8", cursor="crosshair",
        )
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        canvas_vsb = ttk.Scrollbar(canvas_outer, orient=tk.VERTICAL, command=self.canvas.yview)
        canvas_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas_hsb = ttk.Scrollbar(canvas_outer, orient=tk.HORIZONTAL, command=self.canvas.xview)
        canvas_hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.configure(yscrollcommand=canvas_vsb.set, xscrollcommand=canvas_hsb.set)

        self.canvas.bind("<ButtonPress-1>", self._on_canvas_press)
        self.canvas.bind("<B1-Motion>", self._on_canvas_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_canvas_release)
        self.canvas.bind("<Motion>", self._on_canvas_hover)

        result_outer = ttk.LabelFrame(paned, text="추출된 표", padding=2)
        paned.add(result_outer, weight=1)

        self.result_text = ScrolledText(
            result_outer, wrap=tk.NONE,
            font=("Courier", 9), state=tk.DISABLED,
            width=60,
        )
        self.result_text.pack(fill=tk.BOTH, expand=True)

        self.status_var = tk.StringVar(value="PDF를 불러오세요.")
        ttk.Label(root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W).grid(
            row=2, column=0, sticky=tk.EW
        )

    def _browse_pdf(self) -> None:
        p = filedialog.askopenfilename(
            title="PDF 선택",
            filetypes=[("PDF 파일", "*.pdf"), ("모든 파일", "*.*")],
        )
        if p:
            self.pdf_path_var.set(p)

    def _load_pdf(self) -> None:
        path_str = self.pdf_path_var.get().strip()
        if not path_str:
            messagebox.showwarning("경고", "PDF 경로를 입력하세요.")
            return
        path = Path(path_str)
        if not path.is_file():
            messagebox.showerror("오류", f"파일을 찾을 수 없습니다:\n{path}")
            return

        try:
            if self.fitz_doc:
                self.fitz_doc.close()
            if self.plumber_pdf:
                self.plumber_pdf.close()
        except Exception:
            pass

        try:
            self.fitz_doc = fitz.open(str(path))
            self.plumber_pdf = pdfplumber.open(str(path))
        except Exception as e:
            messagebox.showerror("오류", f"PDF 열기 실패:\n{e}")
            return

        self.pdf_path = path
        total = len(self.fitz_doc)

        self._detect_page1_indices()
        if not self.page_indices:
            self.page_indices = list(range(total))
            self.status_var.set(f"Page 1 자동 감지 실패 → 전체 {total}페이지 표시")
        else:
            self.status_var.set(
                f"Page 1 유형 {len(self.page_indices)}페이지 감지 (PDF 전체 {total}페이지)"
            )

        self.cur_pos = 0
        self._show_current_page()

    def _detect_page1_indices(self) -> None:
        indices = []
        for i, page in enumerate(self.plumber_pdf.pages):
            try:
                text = page.extract_text() or ""
                if any(kw in text for kw in _PAGE1_KEYWORDS):
                    indices.append(i)
            except Exception:
                pass
        self.page_indices = indices

    def _prev_page(self) -> None:
        if not self.page_indices:
            return
        self.cur_pos = max(0, self.cur_pos - 1)
        self._show_current_page()

    def _next_page(self) -> None:
        if not self.page_indices:
            return
        self.cur_pos = min(len(self.page_indices) - 1, self.cur_pos + 1)
        self._show_current_page()

    def _current_pdf_idx(self) -> Optional[int]:
        if not self.page_indices:
            return None
        return self.page_indices[self.cur_pos]

    def _show_current_page(self) -> None:
        idx = self._current_pdf_idx()
        if idx is None:
            return

        total = len(self.page_indices)
        self.page_label_var.set(
            f"페이지: {self.cur_pos + 1}/{total}  (PDF#{idx + 1})"
        )

        fpage = self.fitz_doc[idx]
        self.page_w_pdf = fpage.rect.width
        self.page_h_pdf = fpage.rect.height

        pad = self._img_ox
        avail_w = self.CANVAS_MIN_W - pad * 2
        avail_h = self.CANVAS_MIN_H - pad * 2
        self.scale = min(avail_w / self.page_w_pdf, avail_h / self.page_h_pdf)

        self.left_slider.configure(to=self.page_w_pdf)
        self.right_slider.configure(to=self.page_w_pdf)

        mat = fitz.Matrix(self.scale, self.scale)
        pix = fpage.get_pixmap(matrix=mat, alpha=False)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        self._photo = ImageTk.PhotoImage(img)

        disp_w = pix.width + pad * 2
        disp_h = pix.height + pad * 2
        self.canvas.configure(
            scrollregion=(0, 0, disp_w, disp_h),
            width=min(disp_w + 2, self.CANVAS_MIN_W + 2),
        )

        self._collect_pdf_vline_xs(idx)

        self.canvas.delete("all")
        self.canvas.create_image(
            self._img_ox, self._img_oy,
            anchor=tk.NW, image=self._photo, tags="page_image",
        )

        self._extract_and_update()

    def _collect_pdf_vline_xs(self, idx: int) -> None:
        self._pdf_vline_xs = []
        try:
            page = self.plumber_pdf.pages[idx]
            xs = set()
            for edge in (page.edges or []):
                x0 = float(edge.get("x0", 0))
                x1 = float(edge.get("x1", 0))
                if abs(x1 - x0) < 2:
                    xs.add(round((x0 + x1) / 2, 1))
            self._pdf_vline_xs = sorted(xs)
        except Exception:
            pass

    def _extra_lines(self) -> List[float]:
        raw = self.extra_lines_var.get()
        result = []
        for tok in re.split(r"[,\s]+", raw):
            tok = tok.strip()
            if tok:
                try:
                    result.append(float(tok))
                except ValueError:
                    pass
        return result

    def _make_table_settings(self) -> dict:
        left_x = self.left_x_var.get()
        right_x = self.right_x_var.get()
        vlines = sorted({left_x, right_x} | set(self._extra_lines()))
        strategy = self.strategy_var.get()

        base: dict = {
            "explicit_vertical_lines": vlines,
            "snap_tolerance": 2,
            "join_tolerance": 3,
            "edge_min_length": 10,
        }
        if strategy == "lines":
            base["vertical_strategy"] = "lines"
            base["horizontal_strategy"] = "lines"
        elif strategy == "text":
            base["vertical_strategy"] = "text"
            base["horizontal_strategy"] = "text"
        return base

    def _extract_and_update(self, *_) -> None:
        idx = self._current_pdf_idx()
        if idx is None:
            return

        try:
            page = self.plumber_pdf.pages[idx]
            settings = self._make_table_settings()

            tables = page.extract_tables(settings) or []

            table_objs: List[Any] = []
            try:
                table_objs = page.find_tables(settings) or []
            except Exception:
                pass

            self._display_results(tables)
            self._redraw_overlays(table_objs=table_objs)

            n_rows = sum(len(t) for t in tables)
            vlines = self._make_table_settings()["explicit_vertical_lines"]
            self.status_var.set(
                f"표 {len(tables)}개 | {n_rows}행 | "
                f"좌선 X={self.left_x_var.get():.1f} "
                f"우선 X={self.right_x_var.get():.1f} | "
                f"가상선 {vlines} | 전략={self.strategy_var.get()}"
            )
        except Exception as e:
            self.status_var.set(f"추출 오류: {e}")

    def _display_results(self, tables: list) -> None:
        self.result_text.configure(state=tk.NORMAL)
        self.result_text.delete("1.0", tk.END)

        if not tables:
            self.result_text.insert(tk.END, "추출된 표 없음\n")
            self.result_text.configure(state=tk.DISABLED)
            return

        vlines = self._make_table_settings()["explicit_vertical_lines"]
        self.result_text.insert(
            tk.END,
            f"가상 수직선: {[round(x, 1) for x in vlines]}\n"
            f"전략: {self.strategy_var.get()}\n"
            f"{'─' * 60}\n",
        )

        for t_idx, table in enumerate(tables):
            self.result_text.insert(tk.END, f"\n▶ 표 {t_idx + 1}  ({len(table)}행)\n")
            if not table:
                self.result_text.insert(tk.END, "  (비어 있음)\n")
                continue

            n_cols = max((len(row) for row in table), default=0)
            if n_cols == 0:
                continue

            col_w = [0] * n_cols
            for row in table:
                for ci, cell in enumerate(row):
                    if ci >= n_cols:
                        break
                    s = _cell_str(cell)
                    col_w[ci] = max(col_w[ci], min(len(s), 28))

            sep = "+" + "+".join("-" * (w + 2) for w in col_w) + "+"
            self.result_text.insert(tk.END, sep + "\n")

            for row in table:
                parts = []
                for ci in range(n_cols):
                    cell = row[ci] if ci < len(row) else None
                    s = _cell_str(cell)[:28]
                    parts.append(f" {s.ljust(col_w[ci])} ")
                self.result_text.insert(tk.END, "|" + "|".join(parts) + "|\n")
                self.result_text.insert(tk.END, sep + "\n")

        self.result_text.configure(state=tk.DISABLED)

    def _sanitize_sheet_title(self, title: str) -> str:
        title = re.sub(r'[\\/*?:\[\]]', "_", title)
        return title[:31] if title else "Sheet"

    def _extract_tables_for_page(self, idx: int) -> List[list]:
        page = self.plumber_pdf.pages[idx]
        settings = self._make_table_settings()
        return page.extract_tables(settings) or []

    def _write_tables_to_worksheet(self, ws, tables: List[list], page_idx: int) -> None:
        row_ptr = 1
        ws.cell(row=row_ptr, column=1, value=f"PDF Page #{page_idx + 1}")
        row_ptr += 2

        if not tables:
            ws.cell(row=row_ptr, column=1, value="추출된 표 없음")
            return

        for t_idx, table in enumerate(tables, start=1):
            ws.cell(row=row_ptr, column=1, value=f"표 {t_idx}")
            row_ptr += 1

            if not table:
                ws.cell(row=row_ptr, column=1, value="(비어 있음)")
                row_ptr += 2
                continue

            max_cols = max((len(r) for r in table), default=0)

            for row in table:
                for col_idx in range(max_cols):
                    val = row[col_idx] if col_idx < len(row) else ""
                    if val is None:
                        val = ""
                    ws.cell(row=row_ptr, column=col_idx + 1, value=str(val))
                row_ptr += 1

            row_ptr += 2

        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    def _save_current_page_to_excel(self) -> None:
        if not _OPENPYXL_OK:
            messagebox.showerror("오류", "openpyxl이 필요합니다:\npip install openpyxl")
            return

        idx = self._current_pdf_idx()
        if idx is None or self.plumber_pdf is None:
            messagebox.showwarning("경고", "먼저 PDF를 불러오세요.")
            return

        default_name = "extracted_table_current_page.xlsx"
        if self.pdf_path:
            default_name = f"{self.pdf_path.stem}_page{idx + 1}_tables.xlsx"

        save_path = filedialog.asksaveasfilename(
            title="현재 페이지 Excel 저장",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel 파일", "*.xlsx")],
        )
        if not save_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = self._sanitize_sheet_title(f"Page_{idx + 1}")

            tables = self._extract_tables_for_page(idx)
            self._write_tables_to_worksheet(ws, tables, idx)

            wb.save(save_path)
            self.status_var.set(f"Excel 저장 완료: {save_path}")
            messagebox.showinfo("완료", f"현재 페이지 Excel 저장 완료:\n{save_path}")
        except Exception as e:
            messagebox.showerror("오류", f"Excel 저장 실패:\n{e}")

    def _save_all_pages_to_excel(self) -> None:
        if not _OPENPYXL_OK:
            messagebox.showerror("오류", "openpyxl이 필요합니다:\npip install openpyxl")
            return

        if self.plumber_pdf is None or not self.page_indices:
            messagebox.showwarning("경고", "먼저 PDF를 불러오세요.")
            return

        default_name = "extracted_table_all_pages.xlsx"
        if self.pdf_path:
            default_name = f"{self.pdf_path.stem}_all_tables.xlsx"

        save_path = filedialog.asksaveasfilename(
            title="전체 페이지 Excel 저장",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel 파일", "*.xlsx")],
        )
        if not save_path:
            return

        try:
            wb = Workbook()
            first_sheet = True

            for idx in self.page_indices:
                if first_sheet:
                    ws = wb.active
                    first_sheet = False
                else:
                    ws = wb.create_sheet()

                ws.title = self._sanitize_sheet_title(f"Page_{idx + 1}")
                tables = self._extract_tables_for_page(idx)
                self._write_tables_to_worksheet(ws, tables, idx)

            wb.save(save_path)
            self.status_var.set(f"전체 Excel 저장 완료: {save_path}")
            messagebox.showinfo("완료", f"전체 페이지 Excel 저장 완료:\n{save_path}")
        except Exception as e:
            messagebox.showerror("오류", f"전체 Excel 저장 실패:\n{e}")

    def _redraw_overlays(self, *_, table_objs: Optional[List[Any]] = None) -> None:
        self.canvas.delete("overlay")
        if self._photo is None:
            return

        ox, oy = self._img_ox, self._img_oy
        h_c = int(self.page_h_pdf * self.scale)
        sc = self.scale

        if self.show_pdf_lines_var.get():
            idx = self._current_pdf_idx()
            if idx is not None:
                try:
                    page = self.plumber_pdf.pages[idx]
                    for edge in (page.edges or []):
                        ex0 = float(edge.get("x0", 0)) * sc + ox
                        ey0 = float(edge.get("top", 0)) * sc + oy
                        ex1 = float(edge.get("x1", 0)) * sc + ox
                        ey1 = float(edge.get("bottom", 0)) * sc + oy
                        self.canvas.create_line(
                            ex0, ey0, ex1, ey1,
                            fill=COL_EDGE, width=1, tags="overlay",
                        )
                except Exception:
                    pass

        if self.show_cells_var.get() and table_objs:
            for tbl in table_objs:
                try:
                    tx0, tt, tx1, tb = tbl.bbox
                    self.canvas.create_rectangle(
                        tx0 * sc + ox, tt * sc + oy,
                        tx1 * sc + ox, tb * sc + oy,
                        outline=COL_TABLE, width=2, fill="", tags="overlay",
                    )
                except Exception:
                    pass
                try:
                    for row in tbl.rows:
                        for cell_bbox in row.cells:
                            if cell_bbox is None:
                                continue
                            cx0, ctop, cx1, cbot = cell_bbox
                            self.canvas.create_rectangle(
                                cx0 * sc + ox, ctop * sc + oy,
                                cx1 * sc + ox, cbot * sc + oy,
                                outline=COL_CELL, width=1,
                                fill="#fff5e0", stipple="gray12",
                                tags="overlay",
                            )
                except Exception:
                    pass

        for mx in self._extra_lines():
            cx = mx * sc + ox
            self.canvas.create_line(
                cx, oy, cx, oy + h_c,
                fill=COL_MID, width=1, dash=(4, 4), tags="overlay",
            )
            self.canvas.create_text(
                cx + 3, oy + h_c - 14,
                text=f"{mx:.0f}",
                fill=COL_MID, anchor=tk.NW, font=("", 7), tags="overlay",
            )

        lx = self.left_x_var.get() * sc + ox
        self.canvas.create_line(
            lx, oy, lx, oy + h_c,
            fill=COL_LEFT, width=2, dash=(8, 3), tags=("overlay", "left_line"),
        )
        self.canvas.create_text(
            lx + 4, oy + 4,
            text=f"L={self.left_x_var.get():.0f}",
            fill=COL_LEFT, anchor=tk.NW, font=("", 8, "bold"), tags="overlay",
        )

        rx = self.right_x_var.get() * sc + ox
        self.canvas.create_line(
            rx, oy, rx, oy + h_c,
            fill=COL_RIGHT, width=2, dash=(8, 3), tags=("overlay", "right_line"),
        )
        self.canvas.create_text(
            rx - 4, oy + 4,
            text=f"R={self.right_x_var.get():.0f}",
            fill=COL_RIGHT, anchor=tk.NE, font=("", 8, "bold"), tags="overlay",
        )

    def _on_slider_change(self, *_) -> None:
        lv = self.left_x_var.get()
        rv = self.right_x_var.get()
        self._set_entry(self.left_x_entry, f"{lv:.1f}")
        self._set_entry(self.right_x_entry, f"{rv:.1f}")
        self._redraw_overlays()
        self._extract_and_update()

    def _on_left_entry(self, _event=None) -> None:
        try:
            v = float(self.left_x_entry.get())
            v = max(0.0, min(v, self.page_w_pdf))
            self.left_x_var.set(v)
            self._extract_and_update()
        except ValueError:
            pass

    def _on_right_entry(self, _event=None) -> None:
        try:
            v = float(self.right_x_entry.get())
            v = max(0.0, min(v, self.page_w_pdf))
            self.right_x_var.set(v)
            self._extract_and_update()
        except ValueError:
            pass

    @staticmethod
    def _set_entry(entry: ttk.Entry, text: str) -> None:
        if entry.focus_get() == entry:
            return
        entry.delete(0, tk.END)
        entry.insert(0, text)

    def _canvas_x_to_pdf(self, cx: float) -> float:
        return (cx - self._img_ox) / self.scale

    def _clamp_pdf_x(self, x: float) -> float:
        return max(0.0, min(x, self.page_w_pdf))

    def _on_canvas_press(self, event: tk.Event) -> None:
        lx_c = self.left_x_var.get() * self.scale + self._img_ox
        rx_c = self.right_x_var.get() * self.scale + self._img_ox
        if abs(event.x - lx_c) <= DRAG_THRESHOLD_PX:
            self._drag_target = "left"
        elif abs(event.x - rx_c) <= DRAG_THRESHOLD_PX:
            self._drag_target = "right"
        else:
            self._drag_target = None

    def _on_canvas_drag(self, event: tk.Event) -> None:
        if self._drag_target is None:
            return
        pdf_x = self._clamp_pdf_x(self._canvas_x_to_pdf(event.x))
        if self._drag_target == "left":
            self.left_x_var.set(pdf_x)
            self._set_entry(self.left_x_entry, f"{pdf_x:.1f}")
        else:
            self.right_x_var.set(pdf_x)
            self._set_entry(self.right_x_entry, f"{pdf_x:.1f}")
        self._redraw_overlays()

    def _on_canvas_release(self, event: tk.Event) -> None:
        if self._drag_target is not None:
            self._drag_target = None
            self._extract_and_update()

    def _on_canvas_hover(self, event: tk.Event) -> None:
        if self._photo is None:
            return
        lx_c = self.left_x_var.get() * self.scale + self._img_ox
        rx_c = self.right_x_var.get() * self.scale + self._img_ox
        near = (abs(event.x - lx_c) <= DRAG_THRESHOLD_PX
                or abs(event.x - rx_c) <= DRAG_THRESHOLD_PX)
        self.canvas.configure(cursor="sb_h_double_arrow" if near else "crosshair")

    def _snap_to_pdf_line(self, target: str) -> None:
        if not self._pdf_vline_xs:
            messagebox.showinfo("스냅", "현재 페이지에 PDF 수직선 데이터가 없습니다.")
            return
        cur = self.left_x_var.get() if target == "left" else self.right_x_var.get()
        nearest = min(self._pdf_vline_xs, key=lambda x: abs(x - cur))
        if target == "left":
            self.left_x_var.set(nearest)
            self._set_entry(self.left_x_entry, f"{nearest:.1f}")
        else:
            self.right_x_var.set(nearest)
            self._set_entry(self.right_x_entry, f"{nearest:.1f}")
        self._extract_and_update()

    def on_close(self) -> None:
        for doc in (self.fitz_doc, self.plumber_pdf):
            try:
                if doc:
                    doc.close()
            except Exception:
                pass
        self.root.destroy()


def _cell_str(cell: Any) -> str:
    if cell is None:
        return ""
    return str(cell).replace("\r\n", "?").replace("\n", "?").replace("\r", "?")


def main() -> None:
    if not _FITZ_OK:
        print("PyMuPDF(fitz)가 필요합니다: pip install PyMuPDF")
        sys.exit(1)
    if not _PIL_OK:
        print("Pillow가 필요합니다: pip install Pillow")
        sys.exit(1)
    if not _OPENPYXL_OK:
        print("openpyxl이 필요합니다: pip install openpyxl")
        sys.exit(1)

    initial = sys.argv[1] if len(sys.argv) > 1 else None

    root = tk.Tk()
    app = App(root, initial_pdf=initial)
    root.protocol("WM_DELETE_WINDOW", app.on_close)
    root.mainloop()


if __name__ == "__main__":
    main()