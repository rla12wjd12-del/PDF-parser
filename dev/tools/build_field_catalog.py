from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, List

import openpyxl


def _norm(s: str) -> str:
    if s is None:
        return ""
    s = str(s).replace("ㆍ", "·")
    s = " ".join(s.split())
    return s.strip()


def _split_specs(s: str) -> List[str]:
    s = _norm(s)
    if not s:
        return []
    parts: List[str] = []
    for p in s.replace("\n", ",").split(","):
        p = _norm(p)
        if p and p not in parts:
            parts.append(p)
    return parts


def build_catalog(xlsx_path: Path) -> Dict[str, List[str]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for r in range(1, ws.max_row + 1):
        a = _norm(ws.cell(r, 1).value)
        b = _norm(ws.cell(r, 2).value)
        if not a and not b:
            continue
        rows.append((a, b))

    if not rows:
        raise RuntimeError("엑셀에서 읽을 데이터가 없습니다.")

    start_idx = 0
    if ("직무" in rows[0][0]) or ("전문" in rows[0][1]) or ("전문" in rows[0][0]):
        start_idx = 1

    out: Dict[str, List[str]] = {}
    for job, specs in rows[start_idx:]:
        job = _norm(job)
        if not job:
            continue
        out.setdefault(job, [])
        for sp in _split_specs(specs):
            if sp not in out[job]:
                out[job].append(sp)
    return out


def main() -> None:
    root = Path(__file__).resolve().parents[1]

    p = argparse.ArgumentParser(
        description="직무분야/전문분야 엑셀(.xlsx)을 data/field_catalog.json으로 변환"
    )
    p.add_argument(
        "--xlsx",
        default=None,
        help="입력 xlsx 경로(미지정 시: 프로젝트 루트의 첫 번째 .xlsx 사용)",
    )
    p.add_argument(
        "--out",
        default=None,
        help="출력 json 경로(미지정 시: <project_root>/data/field_catalog.json)",
    )
    args = p.parse_args()

    if args.xlsx:
        xlsx_path = Path(args.xlsx)
        if not xlsx_path.is_absolute():
            xlsx_path = (root / xlsx_path).resolve()
    else:
        xlsx_files = list(root.glob("*.xlsx"))
        if not xlsx_files:
            raise SystemExit(
                "입력 xlsx를 찾을 수 없습니다.\n"
                "- 방법1) --xlsx 로 파일 경로를 지정하세요.\n"
                "- 방법2) 프로젝트 루트에 xlsx를 두세요. (예: '직무분야 및 전문분야.xlsx')"
            )
        xlsx_path = xlsx_files[0]

    if not xlsx_path.exists():
        raise SystemExit(f"입력 xlsx 파일이 없습니다: {xlsx_path}")

    catalog = build_catalog(xlsx_path)

    if args.out:
        out_path = Path(args.out)
        if not out_path.is_absolute():
            out_path = (root / out_path).resolve()
    else:
        out_path = root / "data" / "field_catalog.json"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] wrote: {out_path}")
    print(f"[INFO] jobs: {len(catalog)}, specialties: {sum(len(v) for v in catalog.values())}")


if __name__ == "__main__":
    main()

