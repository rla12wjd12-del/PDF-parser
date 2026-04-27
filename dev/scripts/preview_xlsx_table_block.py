#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx에서 [TABLE] 블록(표 덤프)의 상단 일부를 구조적으로 미리보기한다.
토큰이 깨져도 '열 수/비어있지 않은 좌우 컬럼' 등 형태를 확인하는 목적.

사용:
  python scripts/preview_xlsx_table_block.py "excel_output/손인호_L27_R560_lines_p005_p113.xlsx" --sheets p005 p113
"""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl


def _s(v: object, n: int = 24) -> str:
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").replace("\r", "\n").replace("\n", " ")[:n]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--sheets", nargs="*", default=None)
    ap.add_argument("--rows", type=int, default=8, help="블록 시작 후 출력 행 수")
    ap.add_argument("--cols", type=int, default=10, help="출력 열 수")
    args = ap.parse_args()

    p = args.xlsx
    if not p.is_file():
        print("[ERROR] not found")
        return 2

    wb = openpyxl.load_workbook(p, data_only=True)
    targets = args.sheets or wb.sheetnames[:3]

    for name in targets:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        marker = None
        max_row = ws.max_row or 0
        for r in range(1, max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and v.startswith("[TABLE]"):
                marker = r
                break
        print(f"\n=== sheet={name} marker_row={marker} ===")
        if marker is None:
            continue
        for rr in range(marker, min(max_row, marker + args.rows) + 1):
            row = [_s(ws.cell(rr, c).value) for c in range(1, args.cols + 1)]
            non_empty = sum(1 for x in row if x.strip())
            print(f"{rr:04d} non_empty={non_empty:02d} | " + " | ".join(row))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

