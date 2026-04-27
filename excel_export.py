#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
파싱 결과(dict/JSON)를 하나의 Excel(.xlsx) 파일로보냅니다.
최상위 키(섹션)마다 워크북 내 시트를 만듭니다.
"""

from __future__ import annotations

from pathlib import Path
import re
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def json_dumps_compact(obj: Any) -> str:
    import json

    return json.dumps(obj, ensure_ascii=False, separators=(",", ":"))


_RE_INT = re.compile(r"^[+-]?\d+$")
_RE_NUMBER = re.compile(r"^[+-]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?$")
_RE_PERCENT = re.compile(r"^[+-]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?%$")
_RE_DATE_YMD = re.compile(r"^(?P<y>\d{4})[./-](?P<m>\d{1,2})[./-](?P<d>\d{1,2})$")
_RE_DATETIME_YMD_HMS = re.compile(
    r"^(?P<y>\d{4})[./-](?P<m>\d{1,2})[./-](?P<d>\d{1,2})[ T](?P<h>\d{1,2}):(?P<mi>\d{2})(?::(?P<s>\d{2}))?$"
)


def _looks_like_number_text(s: str) -> bool:
    s = s.strip()
    if not s:
        return False
    return bool(_RE_NUMBER.match(s) or _RE_PERCENT.match(s))


def _escape_excel_formula_text(v: Any) -> Any:
    """
    Excel은 셀 값이 '=', '+', '-', '@' 로 시작하면 '수식'으로 해석한다.
    PDF에서 추출된 텍스트가 우연히 '=' 로 시작하는 경우(예: '=3,320m,B=25.0m ...')가 있어
    Excel 복구(removedRecords: sheetX.xml 부분의 수식) 문제가 발생할 수 있다.
    텍스트로 강제하기 위해 선행 apostrophe(')를 붙인다.
    """
    if not isinstance(v, str):
        return v
    if not v:
        return v
    # 음수/양수 숫자 문자열까지 텍스트로 강제하면 숫자 인식이 깨지므로 예외 처리
    if v[0] in ("=", "@"):
        return "'" + v
    if v[0] in ("+", "-") and not _looks_like_number_text(v):
        return "'" + v
    return v


def _coerce_excel_value(v: Any) -> tuple[Any, str | None]:
    """
    문자열로 들어온 숫자/날짜를 Excel이 인식하는 타입으로 변환하고,
    적절한 표시 형식(number_format)을 함께 반환한다.
    """
    if v is None:
        return None, None

    if isinstance(v, (int, float, datetime, date)):
        # 값이 이미 타입으로 들어온 경우: 기본 포맷만 잡아준다.
        if isinstance(v, int):
            return v, "#,##0"
        if isinstance(v, float):
            return v, "#,##0.00"
        if isinstance(v, datetime):
            return v, "yyyy-mm-dd hh:mm"
        if isinstance(v, date):
            return v, "yyyy-mm-dd"
        return v, None

    if isinstance(v, (dict, list)):
        return json_dumps_compact(v), "@"

    if not isinstance(v, str):
        return v, None

    s0 = v
    s = s0.strip()
    if not s:
        return s0, None

    # 수식 인젝션/깨짐 방지: 수식처럼 시작하는 텍스트는 강제 텍스트 처리
    escaped = _escape_excel_formula_text(s0)
    if isinstance(escaped, str) and escaped.startswith("'"):
        return escaped, "@"

    # 날짜/시간 (YYYY-MM-DD, YYYY.MM.DD, YYYY/MM/DD, + optional time)
    m_dt = _RE_DATETIME_YMD_HMS.match(s)
    if m_dt:
        y = int(m_dt.group("y"))
        mo = int(m_dt.group("m"))
        d = int(m_dt.group("d"))
        h = int(m_dt.group("h"))
        mi = int(m_dt.group("mi"))
        sec = int(m_dt.group("s") or 0)
        try:
            return datetime(y, mo, d, h, mi, sec), "yyyy-mm-dd hh:mm"
        except ValueError:
            return s0, None

    m_d = _RE_DATE_YMD.match(s)
    if m_d:
        y = int(m_d.group("y"))
        mo = int(m_d.group("m"))
        d = int(m_d.group("d"))
        try:
            return date(y, mo, d), "yyyy-mm-dd"
        except ValueError:
            return s0, None

    # 퍼센트 (예: 12.5%, 1,234%)
    if _RE_PERCENT.match(s):
        raw = s.replace(",", "").replace("%", "")
        try:
            num = float(raw) / 100.0
        except ValueError:
            return s0, None
        decimals = 0
        if "." in raw:
            decimals = len(raw.split(".", 1)[1])
        fmt = "0%" if decimals == 0 else ("0." + ("0" * decimals) + "%")
        return num, fmt

    # 숫자 (콤마 포함/소수 포함)
    if _RE_NUMBER.match(s):
        raw = s.replace(",", "")
        # 선행 0이 있는 "정수"는 코드/식별자일 수 있으므로 텍스트 유지 (예: 00123)
        if raw.startswith("0") and len(raw) > 1 and "." not in raw and raw.lstrip("0") != "":
            return s0, "@"
        try:
            if "." in raw:
                decimals = len(raw.split(".", 1)[1])
                return float(raw), "#,##0" + ("." + ("0" * decimals) if decimals else "")
            return int(raw), "#,##0"
        except ValueError:
            return s0, None

    return s0, None


def _write_row(ws, values: list[Any]) -> None:
    # openpyxl은 "완전히 빈 시트"도 max_row==1, max_column==1로 보이는 경우가 있어
    # 첫 데이터를 2행부터 쓰게 되며 1행이 빈 행처럼 남을 수 있다.
    if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
        row_idx = 1
    else:
        row_idx = ws.max_row + 1
    for col_idx, v in enumerate(values, start=1):
        coerced, fmt = _coerce_excel_value(v)
        cell = ws.cell(row=row_idx, column=col_idx, value=coerced)
        if fmt:
            cell.number_format = fmt


def _flatten_dict_rows(d: dict[str, Any]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for k, v in d.items():
        if isinstance(v, (dict, list)):
            rows.append([k, json_dumps_compact(v)])
        else:
            rows.append([k, v])
    return rows


def _list_of_dicts_to_rows(items: list[dict[str, Any]]) -> tuple[list[str], list[list[Any]]]:
    cols: list[str] = []
    for row in items:
        for k in row:
            if k not in cols:
                cols.append(k)
    body: list[list[Any]] = []
    for row in items:
        out_row: list[Any] = []
        for c in cols:
            v = row.get(c, "")
            # Excel 셀은 list/dict를 직접 담을 수 없으므로 JSON 문자열로 변환
            if isinstance(v, (dict, list)):
                v = json_dumps_compact(v)
            out_row.append(v)
        body.append(out_row)
    return cols, body


def _write_key_value_sheet(ws, rows: list[list[Any]]) -> None:
    _write_row(ws, ["항목", "값"])
    for r in rows:
        _write_row(ws, [x for x in (r or [])])


def _write_table_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    _write_row(ws, [h for h in (headers or [])])
    for r in rows:
        _write_row(ws, [x for x in (r or [])])


def _autofit_columns(ws, max_width: float = 60.0) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), max_width)


def _safe_sheet_title(name: str, used: set[str]) -> str:
    """Excel 시트 이름 규칙(31자, 금지 문자)에 맞게 조정하고 중복을 피합니다."""
    s = str(name)
    for ch in r"\/*?:[]":
        s = s.replace(ch, "_")
    s = s.strip()[:31] or "Sheet"
    base = s
    n = 1
    while s in used:
        suffix = f"_{n}"
        s = (base[: max(0, 31 - len(suffix))] + suffix).strip() or f"Sheet_{n}"
        n += 1
    used.add(s)
    return s


def _fill_sheet(ws, key: str, value: Any) -> None:
    if isinstance(value, dict):
        rows = _flatten_dict_rows(value)
        if not rows:
            _write_row(ws, ["항목", "값"])
            _write_row(ws, ["(비어 있음)", ""])
        else:
            _write_key_value_sheet(ws, rows)
    elif isinstance(value, list):
        if not value:
            _write_row(ws, ["(데이터 없음)"])
        elif all(isinstance(x, dict) for x in value):
            headers, body = _list_of_dicts_to_rows(value)
            _write_table_sheet(ws, headers, body)
        else:
            _write_row(ws, ["값"])
            for x in value:
                _write_row(ws, [x])
    else:
        _write_row(ws, ["항목", "값"])
        _write_row(ws, [key, value])
    _autofit_columns(ws)


def _write_error_sheet(ws, errors: list[Any]) -> None:
    """파싱 오류 목록을 하이라이트 형식으로 기록한다."""
    from openpyxl.styles import PatternFill, Font

    RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    BOLD_FONT = Font(bold=True)

    if not errors:
        _write_row(ws, ["오류 없음"])
        return

    # 헤더 구성: 오류 목록의 키를 수집
    cols: list[str] = []
    for err in errors:
        if isinstance(err, dict):
            for k in err:
                if k not in cols:
                    cols.append(k)

    if not cols:
        _write_row(ws, ["오류 내용"])
        for e in errors:
            _write_row(ws, [str(e)])
        return

    header_row = ws.max_row + 1 if ws.max_row > 0 else 1
    _write_row(ws, cols)
    for cell in ws[header_row]:
        cell.font = BOLD_FONT
        cell.fill = RED_FILL

    for err in errors:
        if isinstance(err, dict):
            row_vals = []
            for c in cols:
                v = err.get(c, "")
                # 중첩 list/dict는 JSON 문자열로 변환
                if isinstance(v, (list, dict)):
                    import json as _json
                    v = _json.dumps(v, ensure_ascii=False, separators=(",", ":"))
                row_vals.append(v)
            _write_row(ws, row_vals)
        else:
            _write_row(ws, [str(err)] + [""] * (len(cols) - 1))

    _autofit_columns(ws)


def export_dict_to_excel_workbook(data: dict[str, Any], output_path: Path) -> Path:
    """
    하나의 워크북에 섹션(최상위 키)마다 시트를 만들어 저장합니다.
    '_파싱오류' 키가 있으면 오류 전용 시트를 빨간 배경으로 강조합니다.

    Args:
        data: 파싱 결과 전체 dict
        output_path: 저장할 .xlsx 경로

    Returns:
        저장된 파일의 절대 경로
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    used_titles: set[str] = set()
    first = True

    # '_파싱오류' 키는 마지막에 별도 처리하므로 일반 순서에서 제외
    ERROR_KEY = "_파싱오류"

    for key, value in data.items():
        if key == ERROR_KEY:
            continue
        title = _safe_sheet_title(key, used_titles)
        if first:
            ws = wb.active
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title=title)
        _fill_sheet(ws, str(key), value)

    # '_파싱오류' 시트: 항상 마지막에 추가, 오류 없어도 시트 생성
    error_data = data.get(ERROR_KEY, [])
    err_title = _safe_sheet_title("파싱오류", used_titles)
    if first:
        ws_err = wb.active
        ws_err.title = err_title
    else:
        ws_err = wb.create_sheet(title=err_title)
    _write_error_sheet(ws_err, error_data if isinstance(error_data, list) else [])

    wb.save(output_path)
    return output_path.resolve()
